# -*- coding: utf-8 -*-
"""
Created on Tue Sep 02 16:54:20 2014

@author: Kyle Ellefsen
"""
from __future__ import (absolute_import, division,print_function) #, unicode_literals) http://stackoverflow.com/a/28162261/4441139
from future.builtins import (bytes, dict, int, list, object, range, str, ascii, chr, hex, input, next, oct, open, pow, round, super, filter, map, zip)

import numpy as np
import global_vars as g
from process.BaseProcess import BaseProcess, WindowSelector, MissingWindowError, SliderLabel, CheckBox
from PyQt4.QtGui import *
from PyQt4.QtCore import *
from window import Window #to display any 3d array in Flika, just call Window(array_name)
import pyqtgraph as pg
from scipy import ndimage
from analyze.puffs.gaussianFitting import fitGaussian, fitRotGaussian
from process.filters import butterworth_filter
from scipy.signal import filtfilt
import pyqtgraph.opengl as gl
from roi import ROI_rectangle, makeROI
import itertools
import os, inspect,sys
import scipy.stats as st
from scipy import spatial
import matplotlib
from pyqtgraph.dockarea import *
from window import Window
if sys.version_info.major==2:
    import cPickle as pickle # pickle serializes python objects so they can be saved persistantly.  It converts a python object into a savable data structure
else:
    import pickle
import bz2


class OddSlider(SliderLabel):
    def __init__(self,demicals=0):
        SliderLabel.__init__(self,demicals)
    def updateSlider(self,value):
        if value%2==0:
            if value<self.slider.value():
                value-=1
            else:
                value+=1
            self.label.setValue(value)
        self.slider.setValue(int(value*10**self.decimals))
    def updateLabel(self,value):
        if value%2==0:
            value-=1
        self.label.setValue(value)
        
class Threshold_cluster(BaseProcess):
    '''threshold_cluster(binary_window, data_window, highpass_window, roi_width, paddingXY, paddingT_pre, paddingT_post, maxSigmaForGaussianFit, rotatedfit)
    Performs lots of analyses on puffs
    
    Parameters:
        | binary_window (Window) -- Usually the thresholded version of the highpass_window.  If you've already calculated a density image, you can pass that in instead and we will skip the density calculation step
        | data_window (Window) --  This Window is the F/F0 window
        | highpass_window (Window) -- This Window should have the baseline at 0.  It will be used for gaussian fitting and kinetics.    
        | roi_width (int) -- The width of the ROI in pixels
        | paddingXY (int) -- How many pixels do you want to pad the ROI by when fitting with gaussian
        | paddingT_pre (int) -- How many frames before the event detection should we look for the start of the event.
        | paddingT_post (int)  -- How many frames after the event detection should we look for the end of the event.
        | maxSigmaForGaussianFit (int) -- When fitting with a gaussian, what is the upper bound for the sigma (or sigmax and sigmay) parameter 
        | rotatedfit (bool) -- Set this to True to fit to a 2d rotating gaussian.
        | radius (float) -- Puffs seperated by less than this amount (in pixels) will be grouped together in a site.
        | maxPuffLen (int) -- maximum possible duration of puff length in frames.  If puffs last longer, use a bigger value.  This affects 'chunking' when determining distances between pixels.  It's better to overestimate, although increasing too much will slow down analysis.
        | maxPuffDiameter (int) -- This affects 'chunking' when determining distances between pixels.    This value divided by two is also used as the radius for calculating densities
        | density_threshold (float) -- Get rid of all the points you are sure don't belong to puffs.  
        | time_factor (float) -- When looking for the nearest point of higher density, time and space don't need to be weighted equally.  This is how much I will 'shrink' time in order to calculate distances relative to pixels.  Practically, the larger this number, the more points separated in time will tend to be clustered together.  Use a higher time factor if you are sampling at high frame rates.  
        | load_flika_file (bool) -- If this is checked, the flika file with the same name as the data_window file will be loaded, if it exists.
    Returns:
        newWindow
    '''
    def __init__(self):
        super().__init__()
    def gui(self):
        self.gui_reset()
        
        binary_window=WindowSelector()
        data_window=WindowSelector()
        highpass_window=WindowSelector()
        roi_width=OddSlider(0)
        roi_width.setRange(1,21)
        roi_width.setValue(1)
        roi_width.setSingleStep(2)
        paddingXY=SliderLabel(0)
        paddingXY.setRange(0,100)
        paddingXY.setValue(20)
        paddingT_pre=SliderLabel(0); paddingT_pre.setRange(0,100); paddingT_pre.setValue(15)
        paddingT_post=SliderLabel(0); paddingT_post.setRange(0,100); paddingT_post.setValue(15)
        maxSigmaForGaussianFit=SliderLabel(0); maxSigmaForGaussianFit.setRange(0,100); maxSigmaForGaussianFit.setValue(10)
        rotatedfit=CheckBox()
        rotatedfit.setValue(True)
        radius=SliderLabel(2)
        radius.setRange(.01,10)
        radius.setSingleStep(.1)
        radius.setValue(np.sqrt(2))
        maxPuffLen=SliderLabel(0)
        maxPuffLen.setRange(1,100)
        maxPuffLen.setValue(15)
        maxPuffDiameter=SliderLabel(0)
        maxPuffDiameter.setRange(1,100)
        maxPuffDiameter.setValue(10)
        density_threshold=SliderLabel(2)
        density_threshold.setRange(1,6)
        density_threshold.setValue(2.7)
        time_factor=SliderLabel(3)
        time_factor.setRange(0,20)
        time_factor.setValue(1)
        load_flika_file=QCheckBox()
        load_flika_file.setChecked(True)
        
        
        if 'threshold_cluster_settings' in g.m.settings.d.keys():
            varDict=g.m.settings['threshold_cluster_settings']
            for key in varDict.keys():
                eval(key+'.setValue('+str(varDict[key])+')')
        self.items.append({'name':'binary_window','string':'Binary window containing puffs','object':binary_window})
        self.items.append({'name':'data_window','string':'Data window containing F/F0 data','object':data_window})
        self.items.append({'name':'highpass_window','string':'High pass window containing data with baseline at 0','object':highpass_window})
        self.items.append({'name':'roi_width','string':'roi_width','object':roi_width})
        self.items.append({'name':'paddingXY','string':'paddingXY','object':paddingXY})
        self.items.append({'name':'paddingT_pre','string':'paddingT_pre','object':paddingT_pre})
        self.items.append({'name':'paddingT_post','string':'paddingT_post','object':paddingT_post})
        self.items.append({'name':'maxSigmaForGaussianFit','string':'maxSigmaForGaussianFit','object':maxSigmaForGaussianFit})
        self.items.append({'name':'rotatedfit','string':'rotatedfit','object':rotatedfit})
        self.items.append({'name':'radius','string':'radius','object':radius})
        self.items.append({'name':'maxPuffLen','string':'maxPuffLen','object':maxPuffLen})
        self.items.append({'name':'maxPuffDiameter','string':'maxPuffDiameter','object':maxPuffDiameter})
        self.items.append({'name':'density_threshold','string':'density_threshold','object':density_threshold})
        self.items.append({'name':'time_factor','string':'time_factor','object':time_factor})
        self.items.append({'name':'load_flika_file','string':'load_flika_file','object':load_flika_file})
        super().gui()
        self.ui.setGeometry(QRect(676, 55, 1231, 822))
        
        #self.ui.setGeometry(QRect(442, 116, 1345, 960))
    def __call__(self,binary_window, data_window, highpass_window, roi_width=3,paddingXY=20,paddingT_pre=15, paddingT_post=15,maxSigmaForGaussianFit=10, rotatedfit=True, radius=np.sqrt(2), maxPuffLen=15, maxPuffDiameter=10, density_threshold=None, time_factor=1, load_flika_file=True, keepSourceWindow=False):

        
        g.m.statusBar().showMessage('Performing {}...'.format(self.__name__))
        #if binary_window is None or data_window is None:
            #raise(MissingWindowError("You cannot execute '{}' without selecting a window first.".format(self.__name__)))
#            msg='The Threshold and Cluster analyzer requires a binary window as input.  Select a binary window.'
#            g.m.statusBar().showMessage(msg)
#            self.msgBox = QMessageBox()
#            self.msgBox.setText(msg)
#            self.msgBox.show()
#            return
        filename=data_window.filename
        filename=os.path.splitext(filename)[0]+'.flika'
        if load_flika_file and os.path.isfile(filename):
            print('Found persistentInfo file.  Loading previous results')
            with bz2.BZ2File(filename, 'rb') as f:
                persistentInfo=pickle.load(f)
            puffAnalyzer=PuffAnalyzer(data_window,None,highpass_window,None,persistentInfo)
        else:
            udc=dict()
            udc['roi_width']=roi_width
            udc['paddingXY']=paddingXY
            udc['paddingT_pre']=paddingT_pre
            udc['paddingT_post']=paddingT_post
            udc['maxSigmaForGaussianFit']=maxSigmaForGaussianFit
            udc['rotatedfit']=rotatedfit
            udc['radius']=radius #radius -- all puffs this distance away (measured in pixels) from each other will automatically be grouped together into a site
            udc['maxPuffLen']=maxPuffLen
            udc['maxPuffDiameter']=maxPuffDiameter
            if density_threshold is None:
                if 'threshold_cluster_settings' in g.m.settings.d.keys():
                    density_threshold=g.m.settings['threshold_cluster_settings']['density_threshold']
                else:
                    density_threshold=2.7
            udc['density_threshold']=density_threshold
            udc['time_factor']=time_factor
            puffAnalyzer = PuffAnalyzer(data_window,binary_window,highpass_window,udc)
        g.m.windows.append(puffAnalyzer)
        g.m.statusBar().showMessage('Finished with {}.'.format(self.__name__))
        return puffAnalyzer

threshold_cluster=Threshold_cluster()

class PuffAnalyzer(QWidget):
    def __init__(self,data_window,binary_window,highpass_window,udc,persistentInfo=None,parent=None):
        '''udc -- all the user defined constants'''
        super(PuffAnalyzer,self).__init__(parent) ## Create window with ImageView widget
        g.m.puffAnalyzer=self
        self.data_window=data_window
        self.binary_window=binary_window
        self.highpass_window=highpass_window
        self.mt, self.mx,self.my = self.data_window.image.shape
        self.l=None
        if persistentInfo is not None:
            if 'roi_width' not in persistentInfo.udc.keys():
                persistentInfo.udc['roi_width']=3
            self.loadPersistentInfo(persistentInfo)
        else:
            self.udc=udc
            if set(np.unique(binary_window.image.astype(np.int)))!=set([0,1]): #tests if image is not boolean
                self.Densities=binary_window.image
                self.binary_window.close()
            else:
                self.Densities=None
            getDensities(self) #getDensities calls setupUI after it is finished
    
    def loadPersistentInfo(self,persistentInfo):
        self.udc=persistentInfo.udc
        self.clusters=Clusters(None,None,None,self,persistentInfo)
        self.sites=Sites(self)
        for i in np.arange(len(persistentInfo.sites)):
            self.sites.append(Site([puff for puff in g.m.puffAnalyzer.puffs.puffs if puff.starting_idx in persistentInfo.sites[i]]))
        self.trash=Trash(self)
        for i in np.arange(len(persistentInfo.puffs)):
            if persistentInfo.puffs[i]['trashed']:
                puff=[puff for puff in g.m.puffAnalyzer.puffs.puffs if puff.starting_idx==i][0]
                self.trash.append(puff)
                self.puffs.puffs.remove(puff)
        self.setupUI()
        
    def preSetupUI(self):
        self.sites=Sites(self)
        self.trash=Trash(self)
        self.autoGroupSites(self.udc['radius'])
        g.m.settings['threshold_cluster_settings']=self.udc
        g.m.settings.save()
        self.setupUI()
    def setupUI(self):
        self.setWindowTitle('Puff Analyzer - {}'.format(os.path.basename(self.data_window.name)))
        self.setGeometry(QRect(360, 368, 1552, 351))
        
        
        
        if self.l is not None:
            QWidget().setLayout(self.layout())
        self.l=QGridLayout(self)
        self.setLayout(self.l)
        self.area = DockArea()
        self.l.addWidget(self.area)
        self.d1 = Dock("3D Fit", size=(600,300))
        self.d2 = Dock("Measured amplitude over time", size=(600,400))
        self.d3 = Dock("Control Panel", size=(100,400))
        self.d4 = Dock("Scatter Plot", size=(100,400))
        self.area.addDock(self.d1,'left')
        self.area.addDock(self.d2,'right')
        self.area.addDock(self.d3,'right')
        self.area.addDock(self.d4,'below',self.d1)
        
        self.threeD_plot=threeD_plot(self.puffs[0])
        self.d1.addWidget(self.threeD_plot)
        self.threeD_plot_ON=True
        self.trace_plot=pg.PlotWidget()
        self.puffs[0].plot(self.trace_plot)
        self.d2.addWidget(self.trace_plot)

        self.control_panel=QGridLayout()
        self.currentPuff_spinbox =QSpinBox()
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        self.currentPuff_spinbox.valueChanged.connect(self.setCurrPuff)
        self.discardButton=QPushButton('Delete Puff')
        self.discardButton.pressed.connect(self.discard_currPuff)
        self.togglePuffsButton=QPushButton('Toggle Puffs')
        self.togglePuffsButton.pressed.connect(self.togglePuffs)
        self.toggleSitesButton=QPushButton('Toggle Sites')
        self.toggleSitesButton.pressed.connect(self.toggleSites)
        self.toggleTrashButton=QPushButton('Toggle Trash')
        self.toggleTrashButton.pressed.connect(self.toggleTrash)
        self.toggle3DButton=QPushButton('Toggle 3D Plot')
        self.toggle3DButton.pressed.connect(self.toggle3D)
        self.filterButton=QPushButton('Filter')
        self.filterButton.pressed.connect(self.openFilterGUI)
        self.widenButton=QPushButton('Widen all puff durations')
        self.widenButton.pressed.connect(self.widenPuffDurations)
        self.exportButton=QPushButton('Export')
        self.exportButton.pressed.connect(self.export_gui)
        self.savePointsButton=QPushButton('Save points')
        self.savePointsButton.pressed.connect(self.savePoints)
        self.saveButton=QPushButton('Save (.flika)')
        self.saveButton.pressed.connect(self.save)
        self.control_panel.addWidget(self.currentPuff_spinbox,0,0)
        self.control_panel.addWidget(self.discardButton,1,0)
        self.control_panel.addWidget(self.togglePuffsButton,2,0)
        self.control_panel.addWidget(self.toggleSitesButton,3,0)
        self.control_panel.addWidget(self.toggleTrashButton,4,0)
        self.control_panel.addWidget(self.widenButton,5,0)
        self.control_panel.addWidget(self.filterButton,6,0)
        self.control_panel.addWidget(self.exportButton,7,0)
        self.control_panel.addWidget(self.toggle3DButton,8,0)
        self.control_panel.addWidget(self.savePointsButton,9,0)
        self.control_panel.addWidget(self.saveButton,10,0)
        self.control_panelWidget=QWidget()
        self.control_panelWidget.setLayout(self.control_panel)
        self.d3.addWidget(self.control_panelWidget)
        self.puffsVisible=True
        self.sitesVisible=False
        self.trashVisible=False
        self.currentPuff_spinbox.puffidx=-1
        
        
        self.s1=pg.ScatterPlotItem(size=5, pen=pg.mkPen(None)) #PUFFS
        self.data_window.imageview.addItem(self.s1)
        for puff in self.puffs.puffs:
            x=puff.kinetics['x']
            y=puff.kinetics['y']
            self.s1.addPoints(pos=[[x,y]],data=puff, brush=pg.mkBrush(puff.color),pen=pg.mkPen([0,0,0,255]))
        self.s1.sigClicked.connect(self.clicked)
        self.s2=pg.ScatterPlotItem(size=5,  brush=pg.mkBrush(0, 255, 0, 255),pen=pg.mkPen([0,0,0,255])) #SITES
        self.s2.sigClicked.connect(self.clickedSite)
        self.s2.addPoints(pos=[site.pos for site in self.sites],data=self.sites)
        self.s3=pg.ScatterPlotItem(size=5,  brush=pg.mkBrush(0, 255, 255, 255),pen=pg.mkPen([0,0,0,255])) #TRASH
        self.s3.sigClicked.connect(self.clickedTrash)
        self.s4=pg.ScatterPlotItem(size=10, brush=pg.mkBrush(255, 255, 255, 120),pen=pg.mkPen([0,0,0,255]))
        view = pg.GraphicsLayoutWidget()
        self.d4.addWidget(view)
        self.scatter_viewbox=ScatterViewBox()
        self.scatterPlot=view.addPlot(viewBox=self.scatter_viewbox)
        self.scatter_viewbox.drawFinishedSignal.connect(self.lasso_puffs)
        self.scatter_viewbox.EnterPressedSignal.connect(self.reset_scatter_colors)
        self.scatterPlot.addItem(self.s4)   
        self.lastClickedScatterPt = []
        self.s4.sigClicked.connect(self.clickedScatter)
        

        for puff in self.trash:
            x=puff.kinetics['x']
            y=puff.kinetics['y']
            self.s3.addPoints(pos=[[x,y]],data=puff)
        self.siteAnalyzer=None
        self.lastClicked = None
        self.linkTif()
        self.updateScatter()
        self.setCurrPuff(0)
        self.show()
        
    def save(self):
        g.m.statusBar().showMessage('Saving'); print('Saving')
        persistentInfo=PersistentInfo(self)
        filename=self.data_window.filename
        filename=os.path.splitext(filename)[0]+'.flika'
        with bz2.BZ2File(filename, 'w') as f:
            pickle.dump(persistentInfo, f)
        g.m.statusBar().showMessage('Saved Flika file'); print('Saved Flika file')
        
    def savePoints(self):
        g.m.statusBar().showMessage('Saving puff points'); print('Saving puff points')
        puffs=self.puffs.puffs
        pts=[]
        for puff in puffs:
            k=puff.kinetics
            pts.append(np.array([k['t_peak'], k['x'], k['y']]))
        pts=np.array(pts)
        filename=self.data_window.filename
        filename=os.path.splitext(filename)[0]+'_flika_pts.txt'
        np.savetxt(filename,pts)
        g.m.statusBar().showMessage('Saved puff points'); print('Saved puff points')
        
    def clickedScatter(self, plot, points):
        p=points[0]
        starting_idx=p.data()
        curr_idx=[puff.starting_idx for puff in self.puffs].index(starting_idx)
        self.setCurrPuff(curr_idx)
        print("clicked points", curr_idx)
        
    def updateScatter(self,X_axis='sigma',Y_axis='amplitude'):
        pst=PersistentInfo(self)
        starting_indicies=[]
        peak_amps=[]
        stds=[]
        times=[]
        sigmas=[]
        amps=[]
        color=[]
        results=dict()
        for starting_idx in np.arange(len(pst.clusters)):
            if not pst.puffs[starting_idx]['trashed']:
                before=pst.puffs[starting_idx]['kinetics']['before']
                ontimes=list(set(pst.pixel_idxs[pst.clusters[starting_idx]][:,0]))
                ontimes.sort()
                t0=ontimes[0]-before; tf=ontimes[-1]-before
                trace=pst.puffs[starting_idx]['trace'][t0:tf+1]
                peak_amps.append(np.max(trace))
                amps.append(pst.puffs[starting_idx]['kinetics']['amplitude'])
                stds.append(np.std(trace))
                times.append(ontimes[0])
                sigmas.append(pst.puffs[starting_idx]['gaussianParams'][2])
                color.append(pst.puffs[starting_idx]['color'])
                starting_indicies.append(starting_idx)
        results['time']=np.array(times)
        results['amplitude']=np.array(amps)
        results['sigma']=np.array(sigmas)
        
        self.s4.clear()
        n=len(starting_indicies)
        pos=np.array([results[X_axis],results[Y_axis]])
        spots = [{'pos': pos[:,i], 'data': starting_indicies[i], 'brush':pg.mkBrush(color[i])} for i in range(n)]
        self.s4.addPoints(spots)
                
    def lasso_puffs(self):
        pos_array=np.array([[self.s4.data[i][6], self.s4.data[i][0], self.s4.data[i][1]] for i in np.arange(len(self.s4.data))]) #array where each row is [starting_idx, x, y]
        self.s1.clear()
        for puff in self.puffs.puffs:
            x,y=pos_array[np.argwhere(pos_array[:,0]==puff.starting_idx)[0][0],1:]
            if self.scatter_viewbox.currentROI.contains(x,y):
                puff.color=self.scatter_viewbox.scatter_color
            x=puff.kinetics['x']
            y=puff.kinetics['y']
            self.s1.addPoints(pos=[[x,y]],data=puff, brush=pg.mkBrush(puff.color))
        self.updateScatter()
    def reset_scatter_colors(self):
        self.s1.clear()
        for puff in self.puffs.puffs:
            puff.color=(255,0,0,255)
            x=puff.kinetics['x']
            y=puff.kinetics['y']
            self.s1.addPoints(pos=[[x,y]],data=puff, brush=pg.mkBrush(puff.color), pen=pg.mkPen([0,0,0,255]))
        self.updateScatter()
        
        
    def toggle3D(self):
        if self.threeD_plot_ON:
            self.threeD_plot_ON=False
            self.d1.setParent(None)
            self.d1.label.setParent(None)
        else:
            self.threeD_plot_ON=True
            self.threeD_plot.update_puff(self.puffs.getPuff())
            self.d1 = Dock("3D Fit", size=(600,300))
            self.area.addDock(self.d1,'left')
            self.d1.addWidget(self.threeD_plot)
            
    def linkTif(self):
        self.clusterItem=pg.ImageItem(self.puffs.cluster_im[0])
        self.data_window.imageview.view.addItem(self.clusterItem)
        self.clusterItem.setOpacity(.5)
        self.data_window.sigTimeChanged.connect(self.updateTime)        
        puff=self.puffs.getPuff()
        x=np.floor(puff.kinetics['x']); y=np.floor(puff.kinetics['y'])
        roi_width=self.udc['roi_width']
        r=(roi_width-1)/2
        x0=x-r; x1=x+r+1; y0=y-r; y1=y+r+1;
        pts=[(x0,y0),(x0,y1),(x1,y1),(x1,y0),(x0,y0)]
        self.roi=makeROI('rectangle',pts,self.data_window)
        self.data_window.deleteButtonSignal.disconnect(self.roi.deleteCurrentROI)
        self.redTraces=[]
        self.data_window.keyPressSignal.connect(self.keyPressEvent)
        self.roi.plotSignal.connect(self.linkTracefig)
        self.roi.plot()
        
    def linkTracefig(self):
        g.m.currentTrace.finishedDrawingSignal.connect(self.drawRedOverlay) #hopefully this is not already connected
        g.m.currentTrace.p1.scene().sigMouseClicked.connect(self.clickedTrace)
        g.m.currentTrace.keyPressSignal.connect(self.keyPressEvent)
        self.drawRedOverlay()
        
    def updateTime(self,t):
        self.clusterItem.setImage(self.puffs.cluster_im[t])
            
    def closeEvent(self, event):
        self.close()
        event.accept() # let the window close
    def close(self):
        if self.roi in self.data_window.rois:
            self.roi.delete()
        del self.roi
        if g.m.currentTrace is not None:
            for i in np.arange(len(self.redTraces)):
                g.m.currentTrace.p1.removeItem(self.redTraces[i][0])
            g.m.currentTrace.finishedDrawingSignal.disconnect(self.drawRedOverlay)
            g.m.currentTrace.p1.scene().sigMouseClicked.disconnect(self.clickedTrace)
            g.m.currentTrace.keyPressSignal.disconnect(self.keyPressEvent)
        self.data_window.keyPressSignal.disconnect(self.keyPressEvent)
        self.data_window.sigTimeChanged.disconnect(self.updateTime) 
        self.data_window.imageview.view.removeItem(self.clusterItem)
        if self.data_window in g.m.windows:
            if self.sitesVisible:
                self.data_window.imageview.removeItem(self.s2)
            if self.puffsVisible:
                self.data_window.imageview.removeItem(self.s1)
        else:
            del self.data_window
        if self in g.m.windows:
            g.m.windows.remove(self)
        
        
    def autoGroupSites(self,radius=np.sqrt(2)):
        sites=self.sites[:]
        for site in sites:
            self.sites.remove(site) 
        puffs=self.puffs.puffs
        distances=np.zeros((len(puffs),len(puffs)))
        for i in np.arange(len(puffs)):
            x0,y0=(puffs[i].kinetics['x'], puffs[i].kinetics['y'])
            for j in np.arange(len(puffs)):
                x1,y1=(puffs[j].kinetics['x'], puffs[j].kinetics['y'])
                distances[i,j]=np.sqrt((x1-x0)**2+(y1-y0)**2)
        distances=distances<=radius
        sites=[]
        puffsAdded=set()
        for i in np.arange(len(puffs)):
            pfs=set([s[0] for s in np.argwhere(distances[i])])
            if len(pfs.intersection(puffsAdded))==0:
                sites.append(pfs)
                puffsAdded=puffsAdded.union(pfs)
            else:
                old_sites=[sites.index(site) for site in sites if len(site.intersection(pfs))>0]
                if len(old_sites)==1:
                    idx=old_sites[0]
                    sites[idx]=sites[idx].union(pfs)
                    puffsAdded=puffsAdded.union(pfs)
                else: #merge old sites and new puffs to them, delete old sites
                    sites.append(set.union(set.union(*[sites[i] for i in old_sites]),pfs))
                    for index in sorted(old_sites, reverse=True):
                        del sites[index]
        for site in sites:
            pfs=[puffs[idx] for idx in site]
            self.sites.append(Site(pfs))


    def toggleSites(self):
        if self.sitesVisible:
            self.data_window.imageview.removeItem(self.s2)
        else:
            self.data_window.imageview.addItem(self.s2)
        self.sitesVisible=not self.sitesVisible
    def toggleTrash(self):
        if self.trashVisible:
            self.data_window.imageview.removeItem(self.s3)
        else:
            self.data_window.imageview.addItem(self.s3)
        self.trashVisible=not self.trashVisible
    
    def setCurrPuff(self,value,force=False):
        if force is False and self.currentPuff_spinbox.puffidx==value:
            return
        self.currentPuff_spinbox.puffidx=value
        self.puffs.setIndex(value)
        puff=self.puffs[value]
        if self.threeD_plot_ON:
            self.threeD_plot.update_puff(puff)
        self.trace_plot.clear()
        puff.plot(self.trace_plot)
        self.trace_plot.plotItem.autoRange()
        
        rgnbounds=np.array(g.m.currentTrace.region.getRegion())
        rgnbounds+=puff.kinetics['t_peak']-np.mean(rgnbounds)
        g.m.currentTrace.region.setRegion(tuple(rgnbounds))
#        
        if self.lastClicked is not None:
            self.lastClicked.resetPen()
            self.lastClicked.setBrush(self.lastClicked.data().color)
        point=[s for s in self.s1.points() if s.data() is puff][0]
        point.setPen('y', width=2)
        point.setBrush(puff.color)
        # move the roi to that point
        pts=self.roi.getPoints()
        roi_pt=np.array([(pts[0][0]+pts[2][0])/2,(pts[0][1]+pts[2][1])/2])
        puff_pt=np.array([puff.kinetics['x'],puff.kinetics['y']])
        difference=puff_pt-roi_pt
        
        self.roi.path.translate(QPointF(*difference)) #translates the path
        self.roi.pathitem.setPath(self.roi.path) #sets the roi path; moves the yellow box
        self.roi.getPoints()
        trace=self.roi.getTrace()
        roi_index=g.m.currentTrace.get_roi_index(self.roi)
        g.m.currentTrace.update_trace_full(roi_index,trace)
        for roi in self.roi.linkedROIs:
            roi.draw_from_points(self.roi.getPoints())
            roi.getMask()
            trace=roi.getTrace()
            roi_index=g.m.currentTrace.get_roi_index(roi)
            g.m.currentTrace.update_trace_full(roi_index,trace)
        
        
        
        #self.roi.pathitem.setPath(self.roi.path)
        #self.roi.finish_translate()
        #self.roi.draw_from_points(self.roi.getPoints())
        #self.roi.translate_done.emit()
        self.lastClicked = point
        if self.currentPuff_spinbox.value()!=self.puffs.index:
            self.currentPuff_spinbox.setValue(self.puffs.index)
            
        for p in self.lastClickedScatterPt:
            p.resetPen()
        p=[pt for pt in self.s4.points() if pt.data()==puff.starting_idx][0]
        if p is not None:
            p.setPen('y', width=2)
            self.lastClickedScatterPt= [p]


    def clicked(self, plot, points):
        point=points[0]
        puff=point.data()
        self.setCurrPuff(self.puffs.puffs.index(puff))
    def clickedSite(self,plot,points):
        point=points[0]
        site=point.data()
        if self.siteAnalyzer is not None:
            self.siteAnalyzer=None
        self.siteAnalyzer=SiteAnalyzer(site)
    def clickedTrash(self,plot,points):
        point=points[0]
        puff=point.data()
        self.puffs.addPuffs([puff])
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        self.trash.removePuffs([puff])
        self.setCurrPuff(self.puffs.puffs.index(puff))
        
    def keyPressEvent(self,ev):
        nPuffs=len(self.puffs.puffs)
        old_idx=self.puffs.index
        if ev.key() == Qt.Key_Right:
            if old_idx==nPuffs-1:
                new_idx=0
            else:
                new_idx=old_idx+1
            self.setCurrPuff(new_idx)
        elif ev.key() == Qt.Key_Left:
            if old_idx==0:
                new_idx=nPuffs-1
            else:
                new_idx=old_idx-1
            self.setCurrPuff(new_idx)
        elif ev.key()==Qt.Key_Delete:
            if self.roi.mouseIsOver or self.roi.beingDragged:
                puffs=[pt.data() for pt in self.s1.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
                self.discard_puffs(puffs)                
            else: 
                self.discard_currPuff()
        elif ev.key()==Qt.Key_G:
            #Group all events in roi
            sites=[pt.data() for pt in self.s2.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
            if len(sites)==0:
                return
            puffs=list(itertools.chain(*[site.puffs for site in sites]))
            for site in sites:
                self.sites.remove(site) 
            self.sites.append(Site(puffs))
            self.s2.clear()
            for site in self.sites:
                self.s2.addPoints(pos=[site.pos],data=site)
        elif ev.key()==Qt.Key_U:
            sites=[pt.data() for pt in self.s2.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
            if len(sites)==0:
                return
            puffs=list(itertools.chain(*[site.puffs for site in sites]))
            for site in sites:
                self.sites.remove(site) 
            self.sites.extend([Site([p]) for p in puffs])
            self.s2.clear()
            for site in self.sites:
                self.s2.addPoints(pos=[site.pos],data=site)
            
    def discard_currPuff(self):
        puff=self.puffs.getPuff()
        self.discard_puffs([puff])
        
    def discard_puffs(self,puffs):
        self.trash.addPuffs(puffs)
        self.sites.removePuffs(puffs)
        self.puffs.removePuffs(puffs)
        if len(self.puffs.puffs)==0:
            self.close()
            return 
        self.lastClickedScatterPt=[]
        self.updateScatter()
        self.setCurrPuff(self.puffs.index,force=True)
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        # remove points from scatter plot  

    def togglePuffs(self):
        if self.puffsVisible:
            self.data_window.imageview.removeItem(self.s1)
        else:
            self.data_window.imageview.addItem(self.s1)
        self.puffsVisible=not self.puffsVisible
        
    def drawRedOverlay(self):
        puffs=[pt.data() for pt in self.s1.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
        times=[[puff.kinetics['t_start'],puff.kinetics['t_end']+1] for puff in puffs]
        data=g.m.currentTrace.rois[g.m.currentTrace.get_roi_index(self.roi)]['p1trace'].getData()[1]
        
        x=np.array([np.arange(*times[i]) for i in np.arange(len(times))])
        traces=[data[time] for time in x]
        y=np.array(traces)
        for i in np.arange(len(self.redTraces)):
            g.m.currentTrace.p1.removeItem(self.redTraces[i][0])
        self.redTraces=[]
        for i in np.arange(len(x)):
            self.redTraces.append([g.m.currentTrace.p1.plot(x[i],y[i],pen=pg.mkPen('r')),puffs[i]])
        currentPuff=self.puffs.getPuff()
        if currentPuff in puffs:
            idx=puffs.index(currentPuff)
            self.redTraces[idx][0].setPen(pg.mkPen(color='r',width=3))


    def clickedTrace(self,ev):
        self.EEEE=ev
        pos=ev.pos()
        pos=g.m.currentTrace.vb.mapSceneToView(pos)
        t=pos.x()
        puffs=[pt.data() for pt in self.s1.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
        times=[[puff.kinetics['t_start'],puff.kinetics['t_end']] for puff in puffs]
        try:
            index=[(t<=time[1] and t>=time[0]) for time in times].index(True)
        except ValueError:
            return
        puff=puffs[index]
        if ev.button()==1:
            self.setCurrPuff(self.puffs.puffs.index(puff))
    
    def openFilterGUI(self):
        print("This will eventually allow you to filter puffs by things like amplitude and duration, but I haven't implemented it yet")
    
    def widenPuffDurations(self):
        puffs=self.puffs.puffs
        mt=len(self.data_window.image)
        for puff in puffs:
            puff.kinetics['t_start']-=1
            if puff.kinetics['t_start']<0:
                puff.kinetics['t_start']=0
            puff.kinetics['t_end']+=1
            if puff.kinetics['t_end']>=mt:
                puff.kinetics['t_end']=mt-1
            puff.calcRiseFallTimes()
        puff=self.puffs.getPuff()
        self.trace_plot.clear()
        puff.plot(self.trace_plot)
        
    def export_gui(self):
        filename=g.m.settings['filename']
        directory=os.path.dirname(filename)
        if filename is not None:
            filename= QFileDialog.getSaveFileName(g.m, 'Export Puff Info', directory, '*.xlsx')
        else:
            filename= QFileDialog.getSaveFileName(g.m, 'Export Puff Info', '*.xlsx')
        filename=str(filename)
        if filename=='':
            return False
        else:
            self.export(filename)
    def export(self,filename):
        ''' This function saves out all the info about the puffs        
        '''
        from openpyxl import Workbook
        from openpyxl.cell import get_column_letter
        g.m.statusBar().showMessage('Saving {}'.format(os.path.basename(filename)))
        
        workbook = Workbook() 
        sheet = workbook.create_sheet()
        sheet.title="Puff Data"
        header=['Group #','Group x','Group y','No. Events','Max Amp','x','y','t_peak','Amplitude','sigmax','sigmay','angle','r20','r50','r80','r100','f80','f50','f20','f0']
        for j in np.arange(len(header)):
            col = get_column_letter(j+1)
            sheet.cell("{}{}".format(col,1)).value=header[j]
        row=2
        groupN=1
        for site in self.sites:
            r=str(row)
            groupx,groupy=site.pos
            nEvents=len(site.puffs)
            maxAmp=np.max([puff.kinetics['amplitude'] for puff in site.puffs])
            sheet.cell('A'+r).value=groupN
            sheet.cell('B'+r).value=groupx
            sheet.cell('C'+r).value=groupy
            sheet.cell('D'+r).value=nEvents
            sheet.cell('E'+r).value=maxAmp
            for puff in site.puffs:
                r=str(row)
                k=puff.kinetics
                sheet.cell('F'+r).value=k['x']
                sheet.cell('G'+r).value=k['y']
                
                try: sheet.cell('H'+r).value=int(k['t_peak'])
                except ValueError: pass
                sheet.cell('I'+r).value=k['amplitude']
                if 'sigmax' in k.keys():
                    sheet.cell('J'+r).value=k['sigmax']
                    sheet.cell('K'+r).value=k['sigmay']
                    sheet.cell('L'+r).value=k['angle']
                else:
                    sheet.cell('J'+r).value=k['sigma']
                try: sheet.cell('M'+r).value=int(k['r20'])
                except ValueError: pass
                try: sheet.cell('N'+r).value=int(k['r50'])
                except ValueError: pass
                try: sheet.cell('O'+r).value=int(k['r80'])
                except ValueError: pass
                try: sheet.cell('P'+r).value=int(k['t_peak']-k['t_start'])
                except ValueError: pass
                try: sheet.cell('Q'+r).value=int(k['f80'])
                except ValueError: pass
                try: sheet.cell('R'+r).value=int(k['f50'])
                except ValueError: pass
                try: sheet.cell('S'+r).value=int(k['f20'])
                except ValueError: pass
                try: sheet.cell('T'+r).value=int(k['f0'])
                except ValueError: pass
                row+=1
            groupN+=1
            
        sheet = workbook.create_sheet()
        sheet.title="Site traces"
        groupN=1
        for site in self.sites:
            x=int(np.floor(site.pos[0]))
            y=int(np.floor(site.pos[1]))
            roi_width=g.m.puffAnalyzer.udc['roi_width']
            r=(roi_width-1)/2        
            trace=self.data_window.image[:,x-r:x+r+1,y-r:y+r+1]
            trace=np.mean(np.mean(trace,1),1)
            col=get_column_letter(groupN)
            sheet.cell(col+'1').value="Site #{}".format(groupN)
            for i in np.arange(len(trace)):
                sheet.cell(col+str(i+2)).value=trace[i]
            groupN+=1
        
        sheet = workbook.create_sheet()
        sheet.title="Peak aligned Puff Traces"
        siteN=1
        max_peak_idx=np.max([puff.kinetics['t_peak']-puff.kinetics['t_start'] for puff in self.puffs])
        for puff in self.puffs.puffs:
            col=get_column_letter(siteN)
            peak_idx=puff.kinetics['t_peak']-puff.kinetics['t_start']
            for i in np.arange(len(puff.trace)):
                offset=max_peak_idx-peak_idx
                sheet.cell(col+str(offset+i+1)).value=puff.trace[i]
            siteN+=1
        workbook.save(filename)
        g.m.statusBar().showMessage('Successfully saved {}'.format(os.path.basename(filename)))
            

class threeD_plot(gl.GLViewWidget):
    def __init__(self,puff,parent=None):
        super(threeD_plot,self).__init__(parent)
        self.setCameraPosition(distance=150,elevation=30,azimuth=90)
        image=np.copy(puff.gaussianFit)
        mmax=np.max(puff.mean_image)
        image=image/mmax
        self.p1 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        ##    red   = pow(z * colorMap[0] + colorMap[1], colorMap[2])
        ##    green = pow(z * colorMap[3] + colorMap[4], colorMap[5])
        ##    blue  = pow(z * colorMap[6] + colorMap[7], colorMap[8])
        self.p1.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p1.scale(1, 1, 15.0)
        self.p1.translate(-puff.udc['paddingXY'], -puff.udc['paddingXY'], 0)
        self.addItem(self.p1)
        
        image=np.copy(puff.mean_image)
        image=image/mmax
        self.p2 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        self.p2.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p2.scale(1, 1, 15.0)
        self.p2.translate(-puff.udc['paddingXY'], -puff.udc['paddingXY'], 0)
        self.addItem(self.p2)
        
        self.shiftx=int(np.ceil(image.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)
    def update_puff(self,puff):
        self.p1.translate(-self.shiftx,0,0)
        self.p2.translate(self.shiftx,0,0)
        image=np.copy(puff.gaussianFit)
        mmax=np.max(puff.mean_image)
        image=image/mmax
        self.p1.setData(z=image)
        image=np.copy(puff.mean_image)
        image=image/mmax
        self.p2.setData(z=image)
        self.shiftx=int(np.ceil(image.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)
        #self.setMinimumHeight(300)

class Sites(list):
    def __init__(self,puffAnalyzer):
        super(Sites,self).__init__()
        self.puffAnalyzer=puffAnalyzer
    def removePuffs(self,puffs):
        s2=self.puffAnalyzer.s2
        print(puffs[0].starting_idx)
        idxs=[]
        for puff in puffs:
            try:
                idx=[puff in self[i].puffs for i in np.arange(len(self))].index(True)
                idxs.append(idx)
            except ValueError:
                pass
        idxs=np.unique(idxs)
        idxs=list(idxs)
        idxs.sort(reverse=True)
        if len(idxs)>0:
            for i in idxs:
                site=self[i]
                scatterRemovePoints(s2,[iidx for iidx, pt in enumerate(s2.points()) if pt.data()==site])
                site.removePuffs(puffs)
                if len(site.puffs)==0:
                    self.remove(site)
                else:
                    scatterAddPoints(s2,[site.pos],[site])
                    #self.puffAnalyzer.s2.addPoints(pos=[site.pos],data=site)
            
class Site(object):
    def __init__(self,puffs):
        self.puffs=puffs
        self.pos=self.getPos()
    def getPos(self):
        x=np.mean(np.array([p.kinetics['x'] for p in self.puffs]))
        y=np.mean(np.array([p.kinetics['y'] for p in self.puffs]))
        return [x,y]
    def removePuffs(self,puffs):
        for puff in puffs:
            if puff in self.puffs:
                self.puffs.remove(puff)
        if len(self.puffs)==0:
            self.pos=[np.nan,np.nan]
        else:
            self.pos=self.getPos()
        
class Trash(list):
    def __init__(self,puffAnalyzer):
        super(Trash,self).__init__()
        self.puffAnalyzer=puffAnalyzer
    def addPuffs(self,puffs):
        self.extend(puffs)
        pos=[[puff.kinetics['x'],puff.kinetics['y']] for puff in puffs]
        #self.puffAnalyzer.s3.addPoints(pos=pos,data=puffs)
        scatterAddPoints(self.puffAnalyzer.s3,pos,puffs)
    def removePuffs(self,puffs):
        for puff in puffs:
            if puff in self:
                self.remove(puff)
        s=self.puffAnalyzer.s3
        idxs=[]
        for puff in puffs:
            idxs.append([point['data'] for point in s.data].index(puff))
        scatterRemovePoints(self.puffAnalyzer.s3,idxs)

def getDensities(puffAnalyzer):
    if puffAnalyzer.Densities is None: #if we aren't skipping the density calculation step because we loaded in a density window already.
        binary_image=puffAnalyzer.binary_window.image
        mt,mx,my=binary_image.shape
        B=np.zeros(binary_image.shape)
        maxPuffLen=puffAnalyzer.udc['maxPuffLen']# default: 15 
        maxPuffDiameter=puffAnalyzer.udc['maxPuffDiameter']# default: 10
        if maxPuffLen%2==0:
            maxPuffLen+=1
        if maxPuffDiameter%2==0:
            maxPuffDiameter+=1
        mask=np.zeros((maxPuffLen,maxPuffDiameter,maxPuffDiameter))
        center=np.array([(maxPuffLen-1)/2, (maxPuffDiameter-1)/2, (maxPuffDiameter-1)/2]).astype(np.int)
        for i in np.arange(mask.shape[0]):
            for j in np.arange(mask.shape[1]):
                for k in np.arange(mask.shape[2]):
                    if ((i-center[0])**2)/center[0]**2 + ((j-center[1])**2)/center[1]**2 + ((k-center[2])**2)/center[2]**2 <= 1:
                        mask[i,j,k]=1
        B=np.zeros(binary_image.shape)
        pxls=np.array(np.where(binary_image)).T
        percent=0
        for i,pxl in enumerate(pxls):
            if percent<int(100*i/len(pxls)):
                percent=int(100*i/len(pxls))
                print('Calculating Densities  {}%'.format(percent))
            t,x,y=pxl
            try:
                B[t,x,y]=np.count_nonzero(mask*binary_image[t-center[0]:t+center[0]+1,x-center[1]:x+center[1]+1,y-center[2]:y+center[2]+1])
            except ValueError:
                t0=t-center[0]
                tf=t+center[0]+1
                x0=x-center[1]
                xf=x+center[1]+1
                y0=y-center[2]
                yf=y+center[2]+1
                mask2=mask
                if t0<0:
                    mask2=mask2[center[0]-t:,:,:]
                    t0=0
                if x0<0:
                    mask2=mask2[:,center[1]-x:,:]
                    x0=0
                if y0<0:
                    mask2=mask2[:,:,center[2]-y:]
                    y0=0
                if tf>mt-1:
                    mask2=mask2[:-(tf-mt+1),:,:]
                    tf=mt-1
                if xf>mx-1:
                    mask2=mask2[:,:-(xf-mx+1),:]
                    xf=mx-1
                if yf>my-1:
                    mask2=mask2[:,:,:-(yf-my+1)]
                    yf=my-1
                B[t,x,y]=np.count_nonzero(mask2*binary_image[t0:tf,x0:xf,y0:yf])


        '''Here comes the clustering'''
        #from scipy.stats import binom
        #plot(binom.pmf(np.arange(0,nPixels),nPixels,.001)) #this is the probability distribution for the number of True pixels given a number of pixels
        B=np.log(B+1)
        puffAnalyzer.Densities=B
    denseWindow=Window(puffAnalyzer.Densities)
    denseWindow.setWindowTitle('Density Window')
    thresh_slider=QDoubleSpinBox()
    thresh_slider.setValue(puffAnalyzer.udc['density_threshold'])
    thresh_button=QPushButton('Set Threshold')
    denseWindow.layout.addWidget(thresh_slider)
    denseWindow.layout.addWidget(thresh_button)
    def DensityWindowPressed():
        puffAnalyzer.udc['density_threshold']=thresh_slider.value()
        getClusters(puffAnalyzer)
    thresh_button.pressed.connect(DensityWindowPressed)
    
def getMask(nt=5,nx=5,ny=5):
    mask=np.zeros((nt,nx,ny))
    center=np.array([(nt-1)/2, (nx-1)/2, (ny-1)/2]).astype(np.int)
    for i in np.arange(mask.shape[0]):
        for j in np.arange(mask.shape[1]):
            for k in np.arange(mask.shape[2]):
                if ((i-center[0])**2)/center[0]**2 + ((j-center[1])**2)/center[1]**2 + ((k-center[2])**2)/center[2]**2 <= 1:
                    mask[i,j,k]=1
    return mask, center
    
def getClusters(puffAnalyzer):
    puffAnalyzer=g.m.puffAnalyzer
    mt,mx,my=puffAnalyzer.data_window.image.shape
    maxPuffLen=puffAnalyzer.udc['maxPuffLen']
    maxPuffDiameter=puffAnalyzer.udc['maxPuffDiameter']
    density_threshold=puffAnalyzer.udc['density_threshold'] #2.7 #I came up with this threshold by looking at it, but there is a way to emperically determine what probability this corresponds to by using the binomial theorem, the number of voxels in our sphere, and the fact that only .1% of pixels will randomly be True
    B=np.copy(puffAnalyzer.Densities)
    B=B*(B>density_threshold)
    idxs=np.where(B)
    densities=B[idxs]
    densities_jittered=densities+np.arange(len(densities))/(2*np.float(len(densities))) #I do this so no two densities are the same, so each cluster has a peak.
    C=np.zeros(B.shape)
    C_idx=np.zeros(B.shape,dtype=np.int)
    idxs=np.vstack((idxs[0],idxs[1],idxs[2])).T
    C[idxs[:,0],idxs[:,1],idxs[:,2]]=densities_jittered
    C_idx[idxs[:,0],idxs[:,1],idxs[:,2]]=np.arange(len(idxs))
    print("Number of pixels to analyze: {}".format(len(idxs)))
    time_factor=puffAnalyzer.udc['time_factor'] #1 #this is how much I will 'shrink' time in order to calculate distances relative to pixels
    
    def getHigherPoint(ii,mask,center):
        idx=idxs[ii]
        density=densities_jittered[ii]          
        t,x,y=idx
        center=np.copy(center)
        t0=t-center[0]
        tf=t+center[0]+1
        x0=x-center[1]
        xf=x+center[1]+1
        y0=y-center[2]
        yf=y+center[2]+1
        mask2=np.copy(mask)
        if t0<0:
            mask2=mask2[center[0]-t:,:,:]
            center[0]=t
            t0=0
        elif tf>mt-1:
            mask2=mask2[:-(tf-mt+1),:,:]
            tf=mt-1
        if x0<0:
            mask2=mask2[:,center[1]-x:,:]
            center[1]=x
            x0=0
        elif xf>mx-1:
            mask2=mask2[:,:-(xf-mx+1),:]
            xf=mx-1
        if y0<0:
            mask2=mask2[:,:,center[2]-y:]
            center[2]=y
            y0=0
        elif yf>my-1:
            mask2=mask2[:,:,:-(yf-my+1)]
            yf=my-1
            
        positions=np.array(np.where(mask2*C[t0:tf,x0:xf,y0:yf]>density)).astype(float).T-center
        if len(positions)==0:
            return None
        distances=np.sqrt((positions[:,0]/time_factor)**2+positions[:,1]**2+positions[:,2]**2)
        higher_pt=positions[np.argmin(distances)].astype(np.int)+np.array([t0,x0,y0])+center
        higher_pt=C_idx[higher_pt[0],higher_pt[1],higher_pt[2]]
        return [np.min(distances), higher_pt, density]
    higher_pts=np.zeros((len(idxs),3))
    remander=np.arange(len(idxs))
    percentOn=True
    percent=0
    for r in np.arange(5,51,2):
        print('Finding all higher points in radius {}'.format(r))
        mask,center=getMask(r,r,r)
        oldremander=remander
        remander=[]
        for loop_i, ii in enumerate(oldremander):
            if percentOn and percent<int(100*ii/len(oldremander)):
                percent=int(100*ii/len(oldremander))
                print('Calculating Densities  {}%'.format(percent))
            higher_pt=getHigherPoint(ii,mask,center)
            if higher_pt is not None:
                higher_pts[ii]=higher_pt
            else:
                remander.append(ii)
        percentOn=False
    maxDistance=np.sqrt((r/time_factor)**2+2*r**2)
    #for ii in remander:
    #    print(noname)
    #    higher_pts[ii]=[maxDistance, ii, densities_jittered[ii]]
    if len(remander)==1:
        ii=remander[0]
        higher_pts[ii]=[maxDistance, ii, densities_jittered[ii]]
    elif len(remander)>1:
        dens2=densities_jittered[remander]
        idxs2=idxs[remander]
        D=spatial.distance_matrix(idxs2,idxs2)
        for pt in np.arange(len(D)):
            dd=D[pt,:]
            idx=np.argsort(dd)
            ii=1
            while dens2[idx[0]]>dens2[idx[ii]]: # We are searching for the closest point with a higher density
                ii+=1
                if ii==len(dd): #if this is the most dense point, then no point will have a higher density
                    higher_pts[remander[pt]]= [maxDistance, remander[pt], dens2[pt] ] 
                    break
            if ii!=len(dd): #if this is the most dense point, then no point will have a higher density
                higher_pts[remander[pt]]= [   dd[idx[ii]],  remander[idx[ii]],  dens2[idx[ii]]   ]
    puffAnalyzer.clusters=Clusters(higher_pts,idxs,B.shape, puffAnalyzer)


class Point():
    def __init__(self,idx):
        self.children=[]
        self.idx=idx
    def __repr__(self):
        return str(self.idx)
    def getDescendants(self):
        self.descendants=self.children[:]
        for child in self.children:
            self.descendants.extend(child.getDescendants())
        return self.descendants
        
class Clusters():
    def __init__(self,higher_pts,idxs,movieShape,puffAnalyzer,persistentInfo=None):
        self.persistentInfo=persistentInfo
        if persistentInfo is not None:
            self.idxs=persistentInfo.pixel_idxs
            self.movieShape=persistentInfo.movieShape
            self.clusters=persistentInfo.clusters
            self.puffAnalyzer=puffAnalyzer
            self.getPuffs()
        else:
            self.higher_pts=higher_pts
            self.idxs=idxs
            self.movieShape=movieShape
            self.puffAnalyzer=puffAnalyzer
            self.vb=ClusterViewBox()
            self.pw=pg.PlotWidget(viewBox=self.vb)
            higher_pts_tmp=higher_pts[higher_pts[:,0]>1]
            y=[d[0] for d in higher_pts_tmp] #smallest distance to higher point
            x=[d[2] for d in higher_pts_tmp] # density 
            self.pw.plot(x,np.log(y),pen=None, symbolBrush=QBrush(Qt.blue), symbol='o')
            self.pw.plotItem.axes['left']['item'].setLabel('Smallest distance to denser point (natural logarithm)'); self.pw.plotItem.axes['bottom']['item'].setLabel('Density (natural logarithm)')
            self.pw.show()
            self.vb.drawFinishedSignal.connect(self.manuallySelectClusterCenters)
            self.vb.EnterPressedSignal.connect(self.finished)
        
    def getPuffs(self):
        bounds=[]
        standard_deviations=[]
        origins=[]
        for i in np.arange(len(self.clusters)):
            pos=self.idxs[self.clusters[i]]
            bounds.append(np.array([np.min(pos,0),np.max(pos,0)]))
            standard_deviations.append(np.std(pos[:,1:]-np.mean(pos[:,1:],0)))
            origins.append(np.mean(pos,0))
        self.bounds=np.array(bounds)
        self.standard_deviations=np.array(standard_deviations)
        self.origins=np.array(origins)
    
        mt,mx,my=self.movieShape
        try:
            self.cluster_im=np.zeros((mt,mx,my,4),dtype=np.float16)
        except MemoryError:
            print('There is not enough memory to create the image of clusters.')
        cmap=matplotlib.cm.gist_rainbow
        for i in np.arange(len(self.clusters)):
            color=cmap(int(((i%5)*255./6)+np.random.randint(255./12)))
            for j in np.arange(len(self.clusters[i])):
                t,x,y=self.idxs[self.clusters[i][j]]
                self.cluster_im[t,x,y,:]=color
        if self.persistentInfo is None:
            self.puffAnalyzer.puffs=Puffs(self,self.cluster_im,self.puffAnalyzer)
            self.puffAnalyzer.preSetupUI()
        else:
            self.puffAnalyzer.puffs=Puffs(self,self.cluster_im,self.puffAnalyzer,self.persistentInfo)
    def finished(self):
        print('Finished with clusters! Getting puffs')
        self.getPuffs()
        
    def manuallySelectClusterCenters(self):
        self.pw.plotItem.clear()
        centers=[]
        outsideROI=[]
        for i in np.arange(len(self.higher_pts))[self.higher_pts[:,0]>1]:
            y=np.log(self.higher_pts[i][0])#smallest distance to higher point
            x=self.higher_pts[i][2]# density 
            if self.vb.currentROI.contains(x,y):
                centers.append(i)
            else:
                outsideROI.append(i)
        
        higher_pts2=self.higher_pts[:,1].astype(np.int)
        points=[Point(i) for i in np.arange(len(higher_pts2))]
        loop=np.arange(len(higher_pts2))
        loop=np.delete(loop,centers)
        for i in loop:
            if higher_pts2[i]!=i:
                points[higher_pts2[i]].children.append(points[i])
        
        self.clusters=[]
        for center in centers:
            descendants=points[center].getDescendants()
            cluster=[d.idx for d in descendants]
            cluster=np.array(cluster+[center])
            self.clusters.append(cluster)
        
        # This gets rid of clusters that contain very few True pixels
        centers_with_small_cluster=[]
        centers_with_large_cluster=[]
        cluster_sizes=np.array([len(c) for c in self.clusters])
        for i in np.arange(len(self.clusters),0,-1)-1:
            if cluster_sizes[i]<10:  #this constant can be changed
                centers_with_small_cluster.append(centers[i])
                del self.clusters[i]
            else:
                centers_with_large_cluster.append(centers[i])
    
        self.pw.plot(self.higher_pts[centers_with_large_cluster,2],np.log(self.higher_pts[centers_with_large_cluster,0]),pen=None, symbolBrush=QBrush(Qt.green), symbol='o')
        self.pw.plot(self.higher_pts[centers_with_small_cluster,2],np.log(self.higher_pts[centers_with_small_cluster,0]),pen=None, symbolBrush=QBrush(Qt.red), symbol='o')
        self.pw.plot(self.higher_pts[outsideROI,2],np.log(self.higher_pts[outsideROI,0]),pen=None, symbolBrush=QBrush(Qt.blue), symbol='o')
        
        mt,mx,my=self.movieShape
        try:
            self.cluster_im=np.zeros((mt,mx,my,4),dtype=np.float16)
        except MemoryError:
            print('There is not enough memory to create the image of clusters (error in function manuallySelectClusterCenters).')
        cmap=matplotlib.cm.gist_rainbow
        for i in np.arange(len(self.clusters)):
            color=cmap(int(((i%5)*255./6)+np.random.randint(255./12)))
            for j in np.arange(len(self.clusters[i])):
                t,x,y=self.idxs[self.clusters[i][j]]
                self.cluster_im[t,x,y,:]=color
        Window(self.cluster_im)
    

    
    
class Puffs:
    def __init__(self,clusters,cluster_im,puffAnalyzer,persistentInfo=None):#weakfilt,strongfilt,paddingXY,paddingT_pre,paddingT_post,maxSigmaForGaussianFit,rotatedfit):
        self.puffAnalyzer=puffAnalyzer
        self.udc=puffAnalyzer.udc        
        self.puffs=[]
        self.index=0
        self.clusters=clusters
        self.highpass_window=puffAnalyzer.highpass_window
        self.data_window=puffAnalyzer.data_window
        self.cluster_im=cluster_im
        self.puffs=[Puff(i,self.clusters,self,persistentInfo) for i in np.arange(len(self.clusters.clusters))]

    def __getitem__(self, item):
        return self.puffs[item]
    def removeCurrentPuff(self):
        del self.puffs[self.index]
        if self.index==0:
            return self.index
        else:
            self.index-=1
        return self.index
    def getPuff(self):
        return self.puffs[self.index]
    def increment(self):
        self.index+=1
        if len(self.puffs)<self.index+1:
            self.index=0
    def decrement(self):
        self.index-=1
        if self.index<0:
            self.index=len(self.puffs)-1
    def setIndex(self,index):
        self.index=index
        if len(self.puffs)<self.index+1:
            self.index=0
        elif self.index<0:
            self.index=len(self.puffs)-1
    def removePuffs(self,puffs):
        idxs=[]
        for puff in puffs:
            idxs.append([point['data'] for point in self.puffAnalyzer.s1.data].index(puff))
            self.puffs.remove(puff)
        scatterRemovePoints(self.puffAnalyzer.s1,idxs)
        if self.index>=len(self.puffs):
            self.index=len(self.puffs)-1
            
    def addPuffs(self,puffs):
        s=self.puffAnalyzer.s1
        self.puffs.extend(puffs)
        pos=[[puff.kinetics['x'],puff.kinetics['y']] for puff in puffs]
        scatterAddPoints(s,pos,puffs)
        self.puffAnalyzer.updateScatter()
        #s.addPoints(pos=pos,data=puffs)

        
class Puff:
    def __init__(self,starting_idx,clusters,puffs,persistentInfo=None):
        self.starting_idx=starting_idx
        self.clusters=clusters
        self.puffs=puffs
        self.udc=puffs.udc
        self.color=(255,0,0,255)
        self.originalbounds=self.clusters.bounds[starting_idx] # 2x3 array: [[t_min,x_min,y_min],[t_max,x_max,y_max]]
        t0=self.originalbounds[0][0]
        t1=self.originalbounds[1][0]
        x0=self.originalbounds[0][1]-self.udc['paddingXY']
        x1=self.originalbounds[1][1]+self.udc['paddingXY']
        y0=self.originalbounds[0][2]-self.udc['paddingXY']
        y1=self.originalbounds[1][2]+self.udc['paddingXY']
        mt,mx,my=self.puffs.highpass_window.image.shape
        if t0<0: t0=0
        if y0<0: y0=0
        if x0<0: x0=0
        if t1>=mt: t1=mt-1
        if y1>=my: y1=my-1
        if x1>=mx: x1=mx-1
        self.bounds=[(t0,t1),(x0,x1),(y0,y1)]
        
        if persistentInfo is not None:
            puff=persistentInfo.puffs[starting_idx]
            self.trace=puff['trace']
            self.kinetics=puff['kinetics']
            self.gaussianParams=puff['gaussianParams']
            self.mean_image=puff['mean_image']
            self.gaussianFit=puff['gaussianFit']
            try:
                self.color=puff.color #(255,0,0,255)
            except:
                pass
            return None
        self.trace=None
        self.kinetics=dict()
        
        #######################################################################
        #############          FIND (x,y) ORIGIN       ########################
        #######################################################################
        '''
        For debugging, use the following code:
        self=g.m.puffAnalyzer.puffs.getPuff()
        from analyze.puffs.threshold_cluster import *
        [(t0,t1),(x0,x1),(y0,y1)]=self.bounds
        mt,mx,my=self.puffs.highpass_im.shape
        '''
        
        bb=self.bounds
        before=bb[0][0]-self.udc['paddingT_pre']
        after=bb[0][1]+self.udc['paddingT_post']
        if before<0: before=0
        if after>=mt: after=mt-1;
        self.kinetics['before']=before
        self.kinetics['after']=after
        self.sisterPuffs=[] # the length of this list will show how many gaussians to fit 
        for idx,cluster in enumerate(self.clusters.bounds):
            if np.any(np.intersect1d(np.arange(cluster[0,0],cluster[1,0]),np.arange(t0,t1))):
                if np.any(np.intersect1d(np.arange(cluster[0,1],cluster[1,1]),np.arange(x0,x1))):
                    if np.any(np.intersect1d(np.arange(cluster[0,2],cluster[1,2]),np.arange(y0,y1))):
                        if idx != self.starting_idx:
                            self.sisterPuffs.append(idx)
        I=self.puffs.highpass_window.image[bb[0][0]:bb[0][1]+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        I=np.mean(I,0)
        
        def getFitParams(idx):
            xorigin,yorigin=self.clusters.origins[idx,1:]-np.array([x0,y0])
            sigma=self.clusters.standard_deviations[idx]
            x_lower=xorigin-sigma; x_upper=xorigin+sigma; y_lower=yorigin-sigma; y_upper=yorigin+sigma
            amplitude=np.max(I)/2
            sigma=3
            if self.udc['rotatedfit']:
                sigmax=sigma
                sigmay=sigma
                angle=45
                p0=(xorigin,yorigin,sigmax,sigmay,angle,amplitude)
                #                 xorigin                   yorigin             sigmax, sigmay, angle,    amplitude
                fit_bounds = [(x_lower,x_upper), (y_lower,y_upper),  (2,self.udc['maxSigmaForGaussianFit']), (2,self.udc['maxSigmaForGaussianFit']), (0,90),   (0,np.max(I))]
            else:
                p0=(xorigin,yorigin,sigma,amplitude) 
                #                 xorigin                   yorigin            sigma    amplitude
                fit_bounds = [(x_lower,x_upper), (y_lower,y_upper),  (2,self.udc['maxSigmaForGaussianFit']),    (0,np.max(I))] #[(0.0, 2*self.paddingXY), (0, 2*self.paddingXY),(0,10),(0,10),(0,90),(0,5)]
            return p0, fit_bounds
        p0,fit_bounds=getFitParams(self.starting_idx)
        for puff in self.sisterPuffs:
            sister_p0, sister_fit_bounds=getFitParams(puff)
            p0=p0+sister_p0
            fit_bounds=fit_bounds+sister_fit_bounds
        if self.udc['rotatedfit']:
            p, I_fit, I_fit2= fitRotGaussian(I,p0,fit_bounds,nGaussians=1+len(self.sisterPuffs))
            self.mean_image=I
            self.gaussianFit=I_fit2
            p[0]=p[0]+self.bounds[1][0] #Put back in regular coordinate system.  Add back x
            p[1]=p[1]+self.bounds[2][0] #add back y 
            self.gaussianParams=p
            xorigin,yorigin,sigmax,sigmay,angle,amplitude=self.gaussianParams
        else:
            
            p, I_fit, I_fit2= fitGaussian(I,p0,fit_bounds,nGaussians=1+len(self.sisterPuffs))
            self.mean_image=I
            self.gaussianFit=I_fit2
            p[0]=p[0]+self.bounds[1][0] #Put back in regular coordinate system.  Add back x
            p[1]=p[1]+self.bounds[2][0] #add back y 
            self.gaussianParams=p
            xorigin,yorigin,sigma,amplitude=self.gaussianParams

        if self.udc['rotatedfit']:
            self.kinetics['sigmax']=sigmax; self.kinetics['sigmay']=sigmay; self.kinetics['angle']=angle
        else:
            self.kinetics['sigma']=sigma;
        self.kinetics['x']=xorigin; self.kinetics['y']=yorigin;
        #######################################################################
        #############          FIND PEAK       ########################
        #######################################################################
        if amplitude==0:
            I_norm=np.zeros(self.gaussianFit.shape)
            I_norm2=np.zeros(self.gaussianFit.shape)
        else:
            I_norm=I_fit/np.sum(I_fit)
            I_norm2=I_fit2/np.sum(I_fit2)
            

        #I=self.puffs.highpass_im[before:after+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        I=self.puffs.data_window.image[before:after+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        #baseline=np.mean(I[:,I_norm2<.01])     
        trace=np.zeros((len(I)))
        x=int(np.floor(xorigin))-self.bounds[1][0]
        y=int(np.floor(yorigin))-self.bounds[2][0]
        roi_width=self.udc['roi_width']
        r=(roi_width-1)/2        
        for i in np.arange(len(trace)):
            #trace[i]=I[i,x,y]
            trace[i]=np.mean(I[i,x-r:x+r+1,y-r:y+r+1])
            #trace[i]=2*np.sum((I[i]-baseline)*I_norm)+baseline
            #I'm not sure why the 2 is needed, but this seems to always work out to 1:
            #            from analyze.puffs.gaussianFitting import gaussian
            #            x = np.arange(1000,dtype=float)
            #            y = np.arange(1000,dtype=float)
            #            amplitude=2.8
            #            sigma=7
            #            I=gaussian(x[:,None], y[None,:],100,100,sigma,amplitude)
            #            I_fit=gaussian(x[:,None], y[None,:],100,100,sigma,1)
            #            I_norm=I_fit/np.sum(I_fit)
            #            calc_amp=2*np.sum(I*I_norm)
            #            print('Original amp={}  Calculated amp={}'.format(amplitude,calc_amp))
        self.trace=trace
        self.kinetics['t_start']=t0
        self.kinetics['t_end']=t1
        self.calcRiseFallTimes()
        
    def calcRiseFallTimes(self):
        before=self.kinetics['before']
        t_start=self.kinetics['t_start']-before
        t_end=self.kinetics['t_end']-before
        baseline=self.trace[t_start]
        t_peak=np.argmax(self.trace[t_start:t_end+1])+t_start
        f_peak=self.trace[t_peak]
        if baseline>f_peak:
            baseline=self.trace[t_start]
        amplitude=f_peak-baseline
        thresh20=baseline+amplitude*.2
        thresh50=baseline+amplitude*.5
        thresh80=baseline+amplitude*.8
        tmp=np.argwhere(self.trace>thresh20); tmp=tmp[np.logical_and(tmp>=t_start,tmp<=t_peak)]; 
        if len(tmp)==0: r20=np.nan
        else:  r20=tmp[0]-t_start
        tmp=np.argwhere(self.trace>thresh50); tmp=tmp[np.logical_and(tmp>=t_start,tmp<=t_peak)];
        if len(tmp)==0: r50=np.nan
        else:  r50=tmp[0]-t_start
        tmp=np.argwhere(self.trace>thresh80); tmp=tmp[np.logical_and(tmp>=t_start,tmp<=t_peak)]; 
        if len(tmp)==0: r80=np.nan
        else:  r80=tmp[0]-t_start
        tmp=np.argwhere(self.trace<thresh80); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f80=np.nan
        else: f80=tmp[0]-t_peak
        
        tmp=np.argwhere(self.trace<thresh50); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f50=np.nan
        else: f50=tmp[0]-t_peak
        
        tmp=np.argwhere(self.trace<thresh20); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f20=np.nan
        else: f20=tmp[0]-t_peak
        
        tmp=np.argwhere(self.trace<baseline); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f0=np.nan
        else: 
            f0=tmp[0]
            if f0<t_end:
                t_end=f0
            f0=f0-t_peak
            
        self.kinetics['amplitude']=amplitude
        self.kinetics['baseline']=baseline
        #self.kinetics['t_end']=t_end+before
        self.kinetics['r20']=r20
        self.kinetics['r50']=r50
        self.kinetics['r80']=r80
        self.kinetics['f20']=f20
        self.kinetics['f50']=f50
        self.kinetics['f80']=f80
        self.kinetics['f0']=f0
        self.kinetics['t_peak']=t_peak+before

            
    def plot(self,figure=None):
        if figure is None:
            figure=pg.plot()
        k=self.kinetics
        baseline=k['baseline']; amplitude=k['amplitude']
        #thresh20=baseline+amplitude*.2
        #thresh50=baseline+amplitude*.5
        #thresh80=baseline+amplitude*.8
        x=np.arange(len(self.trace))+k['before']
        figure.plot(x,self.trace,pen=pg.mkPen(width=2))
        #figure.plot(x,self.fStrong,pen=pg.mkPen('g'))
        #figure.plot(x,self.fWeak,pen=pg.mkPen('r'))
        self.peakLine=figure.addLine(y=baseline,pen=pg.mkPen('y',style=Qt.DashLine))
        self.baselineLine=figure.addLine(y=baseline+amplitude,pen=pg.mkPen('y',style=Qt.DashLine))
        self.startLine=figure.addLine(x=k['t_start'],pen=pg.mkPen('y',style=Qt.DashLine),movable=True,bounds=(self.kinetics['before'],self.kinetics['t_peak']))
        self.endLine=figure.addLine(x=k['t_end'],pen=pg.mkPen('y',style=Qt.DashLine),movable=True, bounds=(self.kinetics['t_peak'],self.kinetics['after']))
        self.startLine.sigDragged.connect(self.changeStartTime)
        self.endLine.sigDragged.connect(self.changeEndTime)
    def changeStartTime(self,line):
        time=line.value()
        time=int(np.round(time))
        if time!=line.value():
            self.startLine.setValue(time)
        oldstart=self.kinetics['t_start']
        self.kinetics['t_start']=time
        if oldstart!=time:
            self.calcRiseFallTimes()
            self.baselineLine.setValue(self.kinetics['baseline'])
            self.peakLine.setValue(self.kinetics['baseline']+self.kinetics['amplitude'])
            self.endLine.setValue(self.kinetics['t_end'])
            self.puffs.puffAnalyzer.drawRedOverlay()
    def changeEndTime(self,line):
        time=line.value()
        time=int(np.round(time))
        if time!=line.value():
            self.endLine.setValue(time)
        oldend=self.kinetics['t_end']
        if oldend!=time:   
            self.kinetics['t_end']=time
            self.puffs.puffAnalyzer.drawRedOverlay()
    
    
class SiteAnalyzer(QWidget):
    sigTimeChanged=Signal(int)
    def __init__(self,site,parent=None):
        super(SiteAnalyzer,self).__init__(parent) ## Create window with ImageView widget
        #self=QWidget()
        self.site=site
        nPuffs=len(self.site.puffs)
        self.offsets=[puff.kinetics['t_start'] for puff in self.site.puffs]
        
        cmap=matplotlib.cm.gist_rainbow
        self.colors=[cmap(int(i*255./nPuffs)) for i in np.arange(nPuffs)]
        self.colors=[tuple([int(c*255) for c in col]) for col in self.colors]

        self.setGeometry(QRect(1065, 121, 749, 948))
        self.setWindowTitle('Site Analyzer')
        
        
        self.hbox=QGridLayout()
        self.setLayout(self.hbox)
        self.area = DockArea()
        self.hbox.addWidget(self.area)
        self.d1 = Dock("Measured amplitude over time", size=(500,300))
        self.d2 = Dock("Variables", size=(146,400))
        self.d3 = Dock("Event View", size=(500,400))
        self.d4 = Dock("X over time", size=(500,300))
        self.d5 = Dock("Y over time", size=(500,300))
        self.d6 = Dock("Fitted amplitude over time", size=(500,300))
        self.d7 = Dock("Sigma over time", size=(500,300))
        self.d8 = Dock("3D fit",size=(500,400))
        self.d9 = Dock("Puff Data", size=(500,400))        
        self.area.addDock(self.d6, 'left')      ## place d1 at left edge of dock area (it will fill the whole space since there are no other docks yet)
        self.area.addDock(self.d2, 'right', self.d6)     ## place d2 at right edge of dock area
        self.area.addDock(self.d8, 'bottom')
        self.area.addDock(self.d9, 'below', self.d8)
        self.area.addDock(self.d3, 'above', self.d8) ## place d3 at bottom edge of d1
        self.area.addDock(self.d4, 'below', self.d6)    
        self.area.addDock(self.d5, 'below', self.d4)   
        self.area.addDock(self.d7, 'below', self.d5)
        self.area.addDock(self.d1, 'above', self.d6)    
        
        
        self.p1 = pg.PlotWidget(title="Measured amplitude over time")
        self.d1.addWidget(self.p1)
        self.p2 = pg.PlotWidget(title="Fitted amplitude over time")
        self.d6.addWidget(self.p2)
        self.p3 = pg.PlotWidget(title="Fitted X over time")
        self.d4.addWidget(self.p3)
        self.p4 = pg.PlotWidget(title="Fitted Y over time")
        self.d5.addWidget(self.p4)
        self.p5 = pg.PlotWidget(title="Fitted Sigma over time")
        self.d7.addWidget(self.p5)
        self.puff_3D=puff_3D()
        self.d8.addWidget(self.puff_3D)
        self.table=pg.TableWidget()
        self.d9.addWidget(self.table)
        self.formlayout=QFormLayout()
        self.puffCheckBoxes=[QCheckBox() for puff in site.puffs]
        for i in np.arange(nPuffs):
            self.puffCheckBoxes[i].setChecked(True)
            self.puffCheckBoxes[i].stateChanged.connect(self.replot)
            self.formlayout.addRow('Puff '+str(i), self.puffCheckBoxes[i])
        self.alignAverageCheckBox=QCheckBox()
        self.formlayout.addRow('Align Average', self.alignAverageCheckBox)
        self.currentPuff=QComboBox()
        for i in np.arange(nPuffs):
            self.currentPuff.addItem('Puff '+str(i))
        self.currentPuff.currentIndexChanged.connect(self.changeCurrentPuff)
        self.formlayout.addRow('Current Puff', self.currentPuff)
        #self.exportButton=QPushButton("Export")
        #self.exportButton.clicked.connect(self.export)
        #self.formlayout.addWidget(self.exportButton)
        self.formWidget=QWidget()
        self.formWidget.setLayout(self.formlayout)
        self.d2.addWidget(self.formWidget)
        self.imageview=pg.ImageView()
        puff=self.site.puffs[self.currentPuff.currentIndex()]
        self.bounds=np.copy(puff.bounds)
        self.bounds[0,0]=puff.kinetics['before']
        self.bounds[0,1]=puff.kinetics['after']
        
        bb=self.bounds
        self.I=self.site.puffs[0].puffs.highpass_window.image[bb[0][0]:bb[0][1]+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        
        self.imageview.setImage(self.I)
        self.addedItems=[]
        self.sigTimeChanged.connect(self.updateTime)
        self.imageview.timeLine.sigPositionChanged.connect(self.updateindex)
        self.d3.addWidget(self.imageview)
        self.params=[]
        self.I_fits=[]
        for puff in self.site.puffs:
            param, I_fit=self.getParamsOverTime(puff)
            self.params.append(param)
            self.I_fits.append(I_fit)
        self.previousframe=-1
        self.replot()
        self.updatePuffTable()
        self.updateTime(0)
        self.show()
    def updatePuffTable(self):
        idx=self.currentPuff.currentIndex()
        puff=self.site.puffs[idx]
        dt=-(puff.kinetics['t_start']-puff.kinetics['before'])
        
        params=self.params[idx]
        peak_amp=np.max(params[:,3])
        if len(self.params)>1:
            dtype=[('Frame (absolute)', int), ('Frame (relative)',int),('x', float), ('y', float), ('sigma', float), ('amplitude',float), ('x relative to mean (exclusive)', float), 
               ('y relative to mean (exclusive)',float), ('distance to puff center (exclusive)',float), ('peak amplitude', float)]
            other_params=[]
            for i in np.arange(len(self.params)):
                if i != idx:
                    other_params.append(self.params[i])
            other_params=np.vstack(other_params)
            other_x=np.sum(other_params[:,0]*other_params[:,3])/np.sum(other_params[:,3])
            mean_x=np.sum(params[:,0]*params[:,-3])/np.sum(params[:,-3])
            rel_x=mean_x-other_x
            other_y=np.sum(other_params[:,1]*other_params[:,3])/np.sum(other_params[:,3])
            mean_y=np.sum(params[:,1]*params[:,-3])/np.sum(params[:,-3])
            rel_y=mean_y-other_y
            dist=np.sqrt(rel_x**2+rel_y**2)
            
            data=[tuple([i+dt+puff.kinetics['t_start']]+[i+dt]+list(row)+[rel_x,rel_y,dist,peak_amp]) for i, row in enumerate(params)]
        else:
            dtype=[('Frame (absolute)', int), ('Frame (relative)',int),('x', float), ('y', float), ('sigma', float), ('amplitude',float), ('peak amplitude', float)]
            data=[tuple([i+dt+puff.kinetics['t_start']]+[i+dt]+list(row)+[peak_amp]) for i, row in enumerate(params)]
        data = np.array(data, dtype=dtype)
        self.table.setData(data)
        
    def updateindex(self):
        (idx, t) = self.imageview.timeIndex(self.imageview.timeLine)
        t=int(np.ceil(t))
        self.currentIndex = t
        self.sigTimeChanged.emit(t)
    def replot(self):
        self.error_bars=[]
        for p in [self.p1,self.p2,self.p3,self.p4,self.p5]:
            p.clear()
        currentPuff=self.site.puffs[self.currentPuff.currentIndex()]
        x=-(currentPuff.kinetics['t_start']-currentPuff.kinetics['before'])
        self.vLine1=self.p1.addLine(x=x,pen=pg.mkPen('y'))
        self.vLine2=self.p2.addLine(x=x,pen=pg.mkPen('y'))
        #self.vLine3=self.p3.addLine(x=x,pen=pg.mkPen('y'))
        ##self.vLine4=self.p4.addLine(x=x,pen=pg.mkPen('y'))
        self.vLine5=self.p5.addLine(x=x,pen=pg.mkPen('y'))
        for i,puff in enumerate(self.site.puffs):
            if self.puffCheckBoxes[i].isChecked():
                pen=pg.mkPen(pg.mkColor(self.colors[i]))
                x=np.arange(len(puff.trace))+puff.kinetics['before']-self.offsets[i]
                self.p1.plot(x,puff.trace,pen=pen)
                self.p2.plot(x,self.params[i][:,3],pen=pen) #amplitude
                self.p3.plot(x,self.params[i][:,0],pen=pen) # x 
                self.p4.plot(x,self.params[i][:,1],pen=pen) # y
                self.p5.plot(x,self.params[i][:,2],pen=pen) # sigma
    def getParamsOverTime(self,puff):
        print('getting parameters')
        bb=puff.bounds
        puff.udc['rotatedfit']=False
        def getFitParams(idx):
            xorigin,yorigin=puff.clusters.origins[idx,1:]-np.array([bb[1][0],bb[2][0]])
            sigma=puff.clusters.standard_deviations[idx]
            x_lower=xorigin-sigma; x_upper=xorigin+sigma; y_lower=yorigin-sigma; y_upper=yorigin+sigma
            amplitude=np.max(I)/2
            sigma=3
            if puff.udc['rotatedfit']:
                sigmax=sigma
                sigmay=sigma
                angle=45
                p0=(xorigin,yorigin,sigmax,sigmay,angle,amplitude)
                #                 xorigin                   yorigin             sigmax, sigmay, angle,    amplitude
                fit_bounds = [(x_lower,x_upper), (y_lower,y_upper),  (2,puff.udc['maxSigmaForGaussianFit']), (2,puff.udc['maxSigmaForGaussianFit']), (0,90),   (0,np.max(I))]
            else:
                p0=(xorigin,yorigin,sigma,amplitude) 
                #                 xorigin                   yorigin            sigma    amplitude
                fit_bounds = [(x_lower,x_upper), (y_lower,y_upper),  (1.5,puff.udc['maxSigmaForGaussianFit']),    (0,np.max(I))] #[(0.0, 2*self.paddingXY), (0, 2*self.paddingXY),(0,10),(0,10),(0,90),(0,5)]
            return p0, fit_bounds
        [(t0,t1),(x0,x1),(y0,y1)]=puff.bounds
        params=[]
        I=puff.puffs.highpass_window.image[puff.kinetics['before']:puff.kinetics['after']+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        p0,fit_bounds=getFitParams(puff.starting_idx)
        
        puff.sisterPuffs=[] # the length of this list will show how many gaussians to fit 
        for idx,cluster in enumerate(puff.clusters.bounds):
            if np.any(np.intersect1d(np.arange(cluster[0,0],cluster[1,0]),np.arange(t0,t1))):
                if np.any(np.intersect1d(np.arange(cluster[0,1],cluster[1,1]),np.arange(x0,x1))):
                    if np.any(np.intersect1d(np.arange(cluster[0,2],cluster[1,2]),np.arange(y0,y1))):
                        if idx != puff.starting_idx:
                            puff.sisterPuffs.append(idx)
        for puf in puff.sisterPuffs:
            sister_p0, sister_fit_bounds=getFitParams(puf)
            p0=p0+sister_p0
            fit_bounds=fit_bounds+sister_fit_bounds
        Is=[]
        for t in np.arange(len(I)):
            I_t=I[t]
            if puff.udc['rotatedfit']:
                p, I_fit, I_fit2= fitRotGaussian(I_t,p0,fit_bounds,nGaussians=1+len(puff.sisterPuffs))
                p[0]=p[0]+bb[1][0] #Put back in regular coordinate system.  Add back x
                p[1]=p[1]+bb[2][0] #add back y 
                #xorigin,yorigin,sigmax,sigmay,angle,amplitude=p
            else:
                
                p, I_fit, I_fit2= fitGaussian(I_t,p0,fit_bounds,nGaussians=1+len(puff.sisterPuffs))
                p[0]=p[0]+bb[1][0] #Put back in regular coordinate system.  Add back x
                p[1]=p[1]+bb[2][0] #add back y 
                #xorigin,yorigin,sigma,amplitude=p
            params.append(np.array(p))
            Is.append(I_fit2)
        params=np.array(params)
        Is=np.array(Is)
        return params, Is
        
    def changeCurrentPuff(self):
        idx=self.currentPuff.currentIndex()
        puff=self.site.puffs[idx]
        self.bounds=np.copy(puff.bounds)
        self.bounds[0,0]=puff.kinetics['before']
        self.bounds[0,1]=puff.kinetics['after']
        bb=self.bounds
        self.I=self.site.puffs[idx].puffs.highpass_window.image[bb[0][0]:bb[0][1]+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        self.imageview.setImage(self.I)
        self.updatePuffTable()
        #self.params=[self.getParamsOverTime(puff) for puff in self.site.puffs]
        self.updateTime(0)

    def updateTime(self,frame):
        if frame==self.previousframe:
            return
        self.previousframe=frame
        idx=self.currentPuff.currentIndex()
        currentPuff=self.site.puffs[idx]

        image1=np.copy(self.I[frame])
        mmax=np.max(self.I)
        image1=image1/mmax
        image2=np.copy(self.I_fits[idx][frame])
        image2=image2/mmax
        self.puff_3D.update_images(image1,image2)        

        rel_frame=frame-(currentPuff.kinetics['t_start']-currentPuff.kinetics['before'])
        for line in [self.vLine1,self.vLine2,self.vLine5]:#self.vLine3,self.vLine4,self.vLine5]:
            line.setValue(rel_frame)
        for bar in self.error_bars:
            bar.plot.removeItem(bar)
        self.error_bars=[]
        for item in self.addedItems:
            self.imageview.view.removeItem(item)
        self.addedItems=[]
        
        scatter=pg.ScatterPlotItem(size=8, pen=pg.mkPen(None))
        self.imageview.view.addItem(scatter)
        self.addedItems.append(scatter)
        spots=[]
        for i in np.arange(len(self.site.puffs)):
            rel_frame2=frame-((currentPuff.kinetics['t_start']-currentPuff.kinetics['before'])-(self.site.puffs[i].kinetics['t_start']-self.site.puffs[i].kinetics['before']))
            if self.puffCheckBoxes[i].isChecked() and rel_frame2>=0 and rel_frame2<self.params[i].shape[0]:
                centers=self.params[i][:,:2]-self.bounds[1:,0]
                center=centers[rel_frame2,:]
                amps=np.copy(self.params[i][:,-1])
                amps[amps<=0]=.00000001
                amp=amps[rel_frame2]
                std=self.site.puffs[0].clusters.standard_deviations[i]
                color=np.array(self.colors[i])
                color[3]=50
                ### FIRST ADD THE POINT TO THE SCATTER PLOT
                if 0.8/amp<std: #if the amplitude>0 and the error (0.8/SNR) is within the fitting bounds
                    ### FIRST PLOT THE 
                    
                    """ WARNING!!! the amp should really be the signal to noise ratio, so it is only valid when the image has been ratiod by the standard deviation of the noise """
                    sigma=0.8/amp  
                    sigmas=0.8/amps
                    # add error bars to x and y
                    # first add X
                    color[3]=255
                    err= pg.ErrorBarItem(x=np.array([rel_frame]), y=np.array([self.params[i][rel_frame2,0]]), top=np.array([sigma]), bottom=np.array([sigma]), beam=0.5,pen=pg.mkPen(pg.mkColor(tuple(color))))
                    self.p3.addItem(err)
                    err.plot=self.p3
                    self.error_bars.append(err)
                    # second add Y
                    err= pg.ErrorBarItem(x=np.array([rel_frame]), y=np.array([self.params[i][rel_frame2,1]]), top=np.array([sigma]), bottom=np.array([sigma]), beam=0.5,pen=pg.mkPen(pg.mkColor(tuple(color))))
                    self.p4.addItem(err)
                    err.plot=self.p4
                    self.error_bars.append(err)

                    reasonable_fit=sigmas<2*std
                    bounds=QRectF(QPointF(center[0]-sigma,center[1]-sigma),QSizeF(2*sigma,2*sigma))
                    
                    ### plot outer circle
                    pathitem=QGraphicsEllipseItem(bounds)
                    color[-1]=127 #alpha channel
                    pathitem.setPen(pg.mkPen(pg.mkColor(tuple(color))))
                    pathitem.setBrush(pg.mkColor((0,0,0,0)))
                    self.imageview.view.addItem(pathitem)
                    self.addedItems.append(pathitem)

                    ### plot line
                    frame_i=rel_frame2-6
                    if frame_i<0:
                        frame_i=0
                    alpha=255
                    for ii in np.arange(rel_frame2,frame_i,-1)-1:
                        if alpha<=0 or not reasonable_fit[ii]:
                            break
                        color[-1]=alpha #alpha channel
                        alpha=alpha-(255*1./6.)
                        pathitem=QGraphicsPathItem()
                        pathitem.setPen(pg.mkColor(tuple(color)))
                        path=QPainterPath(QPointF(*centers[ii]))
                        path.lineTo(QPointF(*centers[ii+1]))
                        pathitem.setPath(path)
                        self.imageview.view.addItem(pathitem)
                        self.addedItems.append(pathitem)
                    color[3]=255 #make the spot transparent
                spots.append({'pos':center, 'brush':pg.mkBrush(pg.mkColor(tuple(color)))})     
        scatter.addPoints(spots)
        
class puff_3D(gl.GLViewWidget):
    def __init__(self,parent=None):
        super(puff_3D,self).__init__(parent)
        self.setCameraPosition(distance=150,elevation=30,azimuth=90)
        image=np.zeros((10,10))
        self.p1 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        ##    red   = pow(z * colorMap[0] + colorMap[1], colorMap[2])
        ##    green = pow(z * colorMap[3] + colorMap[4], colorMap[5])
        ##    blue  = pow(z * colorMap[6] + colorMap[7], colorMap[8])
        self.p1.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p1.scale(1, 1, 15.0)
        #self.p1.translate(-puff.udc['paddingXY'], -puff.udc['paddingXY'], 0)
        self.addItem(self.p1)
        self.p2 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        self.p2.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p2.scale(1, 1, 15.0)
        #self.p2.translate(-puff.udc['paddingXY'], -puff.udc['paddingXY'], 0)
        self.addItem(self.p2)
        
        self.shiftx=int(np.ceil(image.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)
    def update_images(self,image1,image2):
        self.p1.setData(z=image1)
        self.p2.setData(z=image2)
        self.p1.translate(-self.shiftx,0,0)
        self.p2.translate(self.shiftx,0,0)
        self.shiftx=int(np.ceil(image1.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)



            
        
        



def findextrema(v):
    v_prime=np.concatenate(([0], np.diff(v)))
    zero_crossing=np.concatenate((np.diff(np.sign(v_prime)),[0]))
    maxima=[1 if v < -1 else 0 for v in zero_crossing]
    minima=[1 if v > 1 else 0 for v in zero_crossing]
    if np.sign(v_prime[-1])>0:
        maxima[-1]=1
    else:
        minima[-1]=1;
    if np.sign(v_prime[0])>0:
        maxima[0]=1
    else:
        minima[0]=1
    maxima=np.nonzero(maxima)[0]
    minima=np.nonzero(minima)[0]
    return maxima, minima



def scatterRemovePoints(scatterplot,idxs):
    i2=[i for i in np.arange(len(scatterplot.data)) if i not in idxs]
    points=scatterplot.points()
    points=points[i2]
    spots=[{'pos':points[i].pos(),'data':points[i].data(),'brush':points[i].brush()} for i in np.arange(len(points))]
    #spots=[{'pos':points[i].pos(),'data':points[i].data(),'brush':points[i].brush(),'pen':points[i].pen()} for i in np.arange(len(points))]
    scatterplot.clear()
    scatterplot.addPoints(spots)
def scatterAddPoints(scatterplot,pos,data):
    points=scatterplot.points()
    spots=[{'pos':points[i].pos(),'data':points[i].data()} for i in np.arange(len(points))]
    spots.extend([{'pos':pos[i],'data':data[i]} for i in np.arange(len(pos))])
    scatterplot.clear()
    scatterplot.addPoints(spots)
#    scatterplot.data=np.empty(len(oldData)-numPts,dtype=scatterplot.data.dtype)
#    i2=[i for i in np.arange(len(oldData)) if i not in idxs]
#    scatterplot.data=oldData[i2]
#    scatterplot.prepareGeometryChange()
#    scatterplot.bounds = [None, None]
#    scatterplot.invalidate()
#    scatterplot.updateSpots(oldData)
#    scatterplot.sigPlotChanged.emit(scatterplot)
#    for pt in scatterplot.data:
#        if pt['item'] is not None:
#            pt['item']._data=pt


class ClusterViewBox(pg.ViewBox):
    drawFinishedSignal=Signal()
    EnterPressedSignal=Signal()
    def __init__(self, *args, **kwds):
        pg.ViewBox.__init__(self, *args, **kwds)
        self.currentROI=None
    def keyPressEvent(self,ev):
        if ev.key() == Qt.Key_Enter or ev.key() == Qt.Key_Return:
            self.EnterPressedSignal.emit()
    def mouseDragEvent(self, ev):
        if ev.button() == Qt.LeftButton:
            ev.accept()
            self.ev=ev
            if ev.isStart():
                pt=self.mapSceneToView(self.ev.buttonDownScenePos())
                self.x=pt.x() # this sets x and y to the button down position, not the current position
                self.y=pt.y()
                #print("Drag start x={},y={}".format(self.x,self.y))
                if self.currentROI is not None:
                    self.currentROI.delete()
                self.currentROI=ROI(self,self.x,self.y)
            if ev.isFinish():
                self.currentROI.drawFinished()
                self.drawFinishedSignal.emit()
            else: # if we are in the middle of the drag between starting and finishing
                pt=self.mapSceneToView(self.ev.scenePos())
                self.x=pt.x() # this sets x and y to the button down position, not the current position
                self.y=pt.y()
                #print("Drag continuing x={},y={}".format(self.x,self.y))
                self.currentROI.extend(self.x,self.y)
        else:
            pg.ViewBox.mouseDragEvent(self, ev)
            
class ScatterViewBox(ClusterViewBox):
    def __init__(self, *args, **kwds):
        ClusterViewBox.__init__(self, *args, **kwds)
        self.colorDialog=QColorDialog()
        self.colorDialog.colorSelected.connect(self.colorSelected)
        self.scatter_color=(255,239,5,255)
    def keyPressEvent(self,ev):
        if ev.key() == Qt.Key_Enter or ev.key() == Qt.Key_Return:
            self.EnterPressedSignal.emit()
    def mouseDragEvent(self, ev):
        if ev.button() == Qt.RightButton:
            ev.accept()
            self.ev=ev
            if ev.isStart():
                pt=self.mapSceneToView(self.ev.buttonDownScenePos())
                self.x=pt.x() # this sets x and y to the button down position, not the current position
                self.y=pt.y()
                #print("Drag start x={},y={}".format(self.x,self.y))
                if self.currentROI is not None:
                    self.currentROI.delete()
                self.currentROI=ROI(self,self.x,self.y)
            if ev.isFinish():
                self.currentROI.drawFinished()
                self.drawFinishedSignal.emit()
            else: # if we are in the middle of the drag between starting and finishing
                pt=self.mapSceneToView(self.ev.scenePos())
                self.x=pt.x() # this sets x and y to the button down position, not the current position
                self.y=pt.y()
                #print("Drag continuing x={},y={}".format(self.x,self.y))
                self.currentROI.extend(self.x,self.y)
        else:
            pg.ViewBox.mouseDragEvent(self, ev)
    def mouseClickEvent(self,ev):
        if ev.button() == Qt.RightButton:
            ev.accept()
            self.colorDialog.open() #chage the color
        else:
            pg.ViewBox.mouseClickEvent(self, ev)
    def colorSelected(self, color):
        self.color=color
        if color.isValid():
            self.scatter_color=tuple((np.array(self.color.getRgbF())*255).astype(np.int))
        
            
class ROI(QWidget):
    def __init__(self,viewbox,x,y):
        QWidget.__init__(self)
        self.viewbox=viewbox
        self.path=QPainterPath(QPointF(x,y))
        self.pathitem=QGraphicsPathItem(self.viewbox)
        self.color=Qt.yellow
        self.pathitem.setPen(QPen(self.color))
        self.pathitem.setPath(self.path)
        self.viewbox.addItem(self.pathitem,ignoreBounds=True)
        self.mouseIsOver=False
    def extend(self,x,y):
        self.path.lineTo(QPointF(x,y))
        self.pathitem.setPath(self.path)
    def getPoints(self):
        points=[]
        for i in np.arange(self.path.elementCount()):
            e=self.path.elementAt(i)
            x=e.x; y=e.y
            if len(points)==0 or points[-1]!=(x,y):
                points.append((x,y))
        self.pts=points
        return self.pts
    def drawFinished(self):
        self.path.closeSubpath()
        self.draw_from_points(self.getPoints())
    def contains(self,x,y):
        return self.path.contains(QPointF(x,y))
    def draw_from_points(self,pts):
        self.path=QPainterPath(QPointF(pts[0][0],pts[0][1]))
        for i in np.arange(len(pts)-1)+1:        
            self.path.lineTo(QPointF(pts[i][0],pts[i][1]))
        self.pathitem.setPath(self.path)
    def delete(self):
        self.viewbox.removeItem(self.pathitem)
        
        
class PersistentInfo:
    def __init__(self,puffAnalyzer):
        self.clusters=puffAnalyzer.clusters.clusters # list of arrays of idxs for each cluster
        self.pixel_idxs=puffAnalyzer.clusters.idxs #[t,x,y] for every pixel determined to be in a puff
        self.puffs={puff.starting_idx:
                         {'kinetics':puff.kinetics, 
                         'trace':puff.trace, 
                         'gaussianParams':puff.gaussianParams, 
                         'mean_image':puff.mean_image,
                         'gaussianFit':puff.gaussianFit,
                         'color':puff.color,
                         'trashed':False} for puff in puffAnalyzer.puffs.puffs}
        self.puffs.update({puff.starting_idx:
                        {'kinetics':puff.kinetics, 
                         'trace':puff.trace, 
                         'gaussianParams':puff.gaussianParams, 
                         'mean_image':puff.mean_image,
                         'gaussianFit':puff.gaussianFit,
                         'color':puff.color,
                         'trashed':True}  for puff in puffAnalyzer.trash})
        self.sites=[[puff.starting_idx for puff in site.puffs] for site in puffAnalyzer.sites]
        self.udc=puffAnalyzer.udc
        self.movieShape=puffAnalyzer.clusters.movieShape
