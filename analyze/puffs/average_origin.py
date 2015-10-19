# -*- coding: utf-8 -*-
"""
Created on Tue Sep 02 16:54:20 2014

@author: Kyle Ellefsen
"""
from __future__ import (absolute_import, division,print_function, unicode_literals)
from future.builtins import (bytes, dict, int, list, object, range, str, ascii, chr, hex, input, next, oct, open, pow, round, super, filter, map, zip)

import numpy as np
import global_vars as g
from process.BaseProcess import BaseProcess, WindowSelector, MissingWindowError, SliderLabel, CheckBox
from PyQt4.QtGui import *
from PyQt4.QtCore import *
import pyqtgraph as pg
from scipy import ndimage
from analyze.puffs.gaussianFitting import fitGaussian, fitRotGaussian
from process.filters import butterworth_filter
from scipy.signal import filtfilt
import pyqtgraph.opengl as gl
from roi import ROI_rectangle
import itertools
import os, inspect

        
class Average_origin(BaseProcess):
    '''average_origin(binary_window, data_window, weakfilt, strongfilt, paddingXY, paddingT_pre, paddingT_post, maxSigmaForGaussianFit, rotatedfit)
    Performs lots of analyses on puffs
    
    Parameters:
        | binary_window (Window)
        | data_window (Window)
        | weakfilt (float) -- The normalized cuttoff frequency of the weak (red) low pass filter
        | strongfilt (float) -- The normalized cuttoff frequency of the strong (green) low pass filter
        | paddingXY (int) -- How many pixels do you want to pad the ROI by when fitting with gaussian
        | paddingT_pre (int) -- How many frames before the event detection should we look for the start of the event.
        | paddingT_post (int)  -- How many frames after the event detection should we look for the end of the event.
        | maxSigmaForGaussianFit (int) -- When fitting with a gaussian, what is the upper bound for the sigma (or sigmax and sigmay) parameter 
        | rotatedfit (bool) -- Set this to True to fit to a 2d rotating gaussian.
        | radius (float) -- Puffs seperated by less than this amount (in pixels) will be grouped together in a site.
    Returns:
        newWindow
    '''
    def __init__(self):
        super().__init__()
    def gui(self):
        self.gui_reset()
        binary_window=WindowSelector()
        data_window=WindowSelector()
        weakfilt=SliderLabel(3)
        weakfilt.setRange(0,1)
        weakfilt.setValue(.7)
        weakfilt.setSingleStep(.01)
        strongfilt=SliderLabel(3)
        strongfilt.setRange(0,1)
        strongfilt.setValue(.5)
        strongfilt.setSingleStep(.01)
        weakfilt.valueChanged.connect(lambda weak: strongfilt.setMaximum(weak))
        strongfilt.valueChanged.connect(lambda strong: weakfilt.setMinimum(strong))
        paddingXY=SliderLabel(0)
        paddingXY.setRange(0,100)
        paddingXY.setValue(20)
        paddingT_pre=SliderLabel(0); paddingT_pre.setRange(0,100); paddingT_pre.setValue(15)
        paddingT_post=SliderLabel(0); paddingT_post.setRange(0,100); paddingT_post.setValue(15)
        maxSigmaForGaussianFit=SliderLabel(0); maxSigmaForGaussianFit.setRange(0,100); maxSigmaForGaussianFit.setValue(10)
        rotatedfit=CheckBox()
        rotatedfit.setValue(True)
        radius=SliderLabel(2)
        radius.setRange(.01,10)
        radius.setSingleStep(.1)
        radius.setValue(np.sqrt(2))
        if 'average_origin_settings' in g.m.settings.d.keys():
            varDict=g.m.settings['average_origin_settings']
            for key in varDict.keys():
                eval(key+'.setValue('+str(varDict[key])+')')
        self.items.append({'name':'binary_window','string':'Binary window containing puffs','object':binary_window})
        self.items.append({'name':'data_window','string':'Data window containing raw data','object':data_window})
        self.items.append({'name':'weakfilt','string':'Weak Low Pass Filter Cutoff Frequency (red)','object':weakfilt})
        self.items.append({'name':'strongfilt','string':'Strong Low Pass Filter Cutoff Frequency (green)','object':strongfilt})
        self.items.append({'name':'paddingXY','string':'paddingXY','object':paddingXY})
        self.items.append({'name':'paddingT_pre','string':'paddingT_pre','object':paddingT_pre})
        self.items.append({'name':'paddingT_post','string':'paddingT_post','object':paddingT_post})
        self.items.append({'name':'maxSigmaForGaussianFit','string':'maxSigmaForGaussianFit','object':maxSigmaForGaussianFit})
        self.items.append({'name':'rotatedfit','string':'rotatedfit','object':rotatedfit})
        self.items.append({'name':'radius','string':'radius','object':radius})
        
        super().gui()
    def __call__(self,binary_window, data_window, weakfilt=.7,strongfilt=.5,paddingXY=20,paddingT_pre=15, paddingT_post=15,maxSigmaForGaussianFit=10, rotatedfit=True, radius=np.sqrt(2), keepSourceWindow=False):

        
        g.m.statusBar().showMessage('Performing {}...'.format(self.__name__))
        if binary_window is None or data_window is None:
            raise(MissingWindowError("You cannot execute '{}' without selecting a window first.".format(self.__name__)))
        if set(np.unique(binary_window.image.astype(np.int)))!=set([0,1]): #tests if image is not boolean
            msg='The Average origin analyzer requires a binary window as input.  Select a binary window.'
            g.m.statusBar().showMessage(msg)
            self.msgBox = QMessageBox()
            self.msgBox.setText(msg)
            self.msgBox.show()
            return
            
        puffs=getPuffs(binary_window.image,data_window.image,weakfilt,strongfilt,paddingXY,paddingT_pre,paddingT_post,maxSigmaForGaussianFit, rotatedfit)
        puffAnalyzer = PuffAnalyzer(puffs,data_window,binary_window,radius)
        #puffAnalyzer.linkTif(data_window)
        g.m.windows.append(puffAnalyzer)
        g.m.statusBar().showMessage('Finished with {}.'.format(self.__name__))
        
        #save parameters
        varnames=[i for i in inspect.getargspec(self.__call__)[0] if i!='self' and i!='keepSourceWindow' and i!='binary_window' and i!='data_window']
        varDict={}
        for name in varnames:
            varDict[name]=eval(name)
        g.m.settings['average_origin_settings']=varDict
        return puffAnalyzer

average_origin=Average_origin()

class PuffAnalyzer(QWidget):
    def __init__(self,puffs,data_window,binary_window,radius=2,parent=None):
        '''radius -- all puffs this distance away (measured in pixels) from each other will automatically be grouped together into a site'''
        super(PuffAnalyzer,self).__init__(parent) ## Create window with ImageView widget
        
        self.puffs=puffs
        self.puffs.puffAnalyzer=self
        self.data_window=data_window
        self.binary_window=binary_window
        self.sites=Sites(self)
        self.trash=Trash(self)
        self.setWindowTitle('Puff Analyzer - {}'.format(os.path.basename(data_window.name)))
        self.setGeometry(QRect(360, 368, 1552, 351))
        self.l=QGridLayout()
        self.threeD_plot=threeD_plot(self.puffs[0])
        self.threeD_plot.setFixedHeight(300)
        self.threeD_plot.setFixedWidth(600)
        self.l_bottom=QGridLayout()
        self.l_bottom.addWidget(self.threeD_plot,0,0)
        self.trace_plot=pg.PlotWidget()
        self.puffs[0].plot(self.trace_plot)
        self.l_bottom.addWidget(self.trace_plot,0,1)        
        self.setLayout(self.l)
        self.s1=pg.ScatterPlotItem(size=5, pen=pg.mkPen(None), brush=pg.mkBrush(0, 0, 255, 255))
        data_window.imageview.addItem(self.s1)
        for puff in self.puffs.puffs:
            x=puff.kinetics['x']
            y=puff.kinetics['y']
            self.s1.addPoints(pos=[[x,y]],data=puff)
        self.lastClicked = None
        self.s1.sigClicked.connect(self.clicked)
        self.s2=pg.ScatterPlotItem(size=5, pen=pg.mkPen(None), brush=pg.mkBrush(0, 255, 0, 255))
        self.s3=pg.ScatterPlotItem(size=5, pen=pg.mkPen(None), brush=pg.mkBrush(0, 255, 255, 255))
        self.s3.sigClicked.connect(self.clickedTrash)
        self.autoGroupSites(radius)
        self.l_top=QGridLayout()
        self.currentPuff_spinbox =QSpinBox()
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        self.currentPuff_spinbox.valueChanged.connect(self.setCurrPuff)
        self.discardButton=QPushButton('Remove Puff')
        self.discardButton.pressed.connect(self.discard_currPuff)
        self.togglePuffsButton=QPushButton('Toggle Puffs')
        self.togglePuffsButton.pressed.connect(self.togglePuffs)
        self.toggleSitesButton=QPushButton('Toggle Sites')
        self.toggleSitesButton.pressed.connect(self.toggleSites)
        self.toggleTrashButton=QPushButton('Toggle Trash')
        self.toggleTrashButton.pressed.connect(self.toggleTrash)
        self.filterButton=QPushButton('Filter')
        self.filterButton.pressed.connect(self.openFilterGUI)
        self.l_top.addWidget(self.currentPuff_spinbox,0,0)
        self.l_top.addWidget(self.discardButton,0,1)
        self.l_top.addWidget(self.togglePuffsButton,0,2)
        self.l_top.addWidget(self.toggleSitesButton,0,3)
        self.l_top.addWidget(self.toggleTrashButton,0,4)
        self.l_top.addWidget(self.filterButton,0,5)
        self.exportButton=QPushButton('Export')
        self.exportButton.pressed.connect(self.export_gui)
        self.l_top.addWidget(self.exportButton,0,6)
        self.l.addLayout(self.l_bottom,1,0)
        self.l.addLayout(self.l_top,0,0)
        self.puffsVisible=True
        self.sitesVisible=False
        self.trashVisible=False
        self.linkTif()
        self.currentPuff_spinbox.puffidx=-1
        self.setCurrPuff(0)
        self.show()
    def linkTif(self):
        tif=self.data_window
        self.roi=ROI_rectangle(tif,0,0)
        tif.currentROI=self.roi
        tif.rois.append(self.roi)
        puff=self.puffs.getPuff()
        (t0,t1),(x0,x1),(y0,y1)=tuple(puff.bounds)
        x=puff.kinetics['x']; y=puff.kinetics['y']
        x0=x-2; x1=x+2; y0=y-2; y1=y+2;
        self.roi.draw_from_points([(x0,y0),(x0,y1),(x1,y1),(x1,y0),(x0,y0)])
        tif.setIndex(np.mean([t0,t1]))
        self.roi.plot()
        tif.deleteButtonSignal.disconnect(self.roi.deleteCurrentROI)
        self.redTraces=[]
        g.m.currentTrace.finishedDrawingSignal.connect(self.drawRedOverlay)
        g.m.currentTrace.p1.scene().sigMouseClicked.connect(self.clickedTrace)
        g.m.currentTrace.keyPressSignal.connect(self.keyPressEvent)
        self.data_window.keyPressSignal.connect(self.keyPressEvent)
        self.drawRedOverlay()
    def closeEvent(self, event):
        if self.roi in self.data_window.rois:
            self.roi.delete()
        del self.roi
        if g.m.currentTrace is not None:
            for i in np.arange(len(self.redTraces)):
                g.m.currentTracep1.removeItem(self.redTraces[i][0])
            g.m.currentTrace.finishedDrawingSignal.disconnect(self.drawRedOverlay)
            g.m.currentTrace.p1.scene().sigMouseClicked.disconnect(self.clickedTrace)
            g.m.currentTrace.keyPressSignal.disconnect(self.keyPressEvent)
        self.data_window.keyPressSignal.disconnect(self.keyPressEvent)
        if self.data_window in g.m.windows:
            if self.sitesVisible:
                self.data_window.imageview.removeItem(self.s2)
            if self.puffsVisible:
                self.data_window.imageview.removeItem(self.s1)
        else:
            del self.data_window
        if self in g.m.windows:
            g.m.windows.remove(self)
        event.accept() # let the window close
        
        
    def autoGroupSites(self,radius=np.sqrt(2)):
        sites=self.sites[:]
        for site in sites:
            self.sites.remove(site) 
        puffs=self.puffs.puffs
        distances=np.zeros((len(puffs),len(puffs)))
        for i in np.arange(len(puffs)):
            x0,y0=(puffs[i].kinetics['x'], puffs[i].kinetics['y'])
            for j in np.arange(len(puffs)):
                x1,y1=(puffs[j].kinetics['x'], puffs[j].kinetics['y'])
                distances[i,j]=np.sqrt((x1-x0)**2+(y1-y0)**2)
        distances=distances<=radius
        sites=[]
        sitesAdded=set()
        for i in np.arange(len(puffs)):
            p=set([s[0] for s in np.argwhere(distances[i])])
            if len(p.intersection(sitesAdded))==0:
                sites.append(p)
                sitesAdded=sitesAdded.union(p)
            else:
                idx=[sites.index(site) for site in sites if len(site.intersection(p))>0][0]
                sites[idx]=sites[idx].union(p)
                sitesAdded=sitesAdded.union(p)
        for site in sites:
            p=[puffs[idx] for idx in site]
            self.sites.append(Site(p))
        self.s2.clear()
        pos=[site.pos for site in self.sites]
        self.s2.addPoints(pos=pos,data=self.sites)

    def toggleSites(self):
        if self.sitesVisible:
            self.data_window.imageview.removeItem(self.s2)
        else:
            self.data_window.imageview.addItem(self.s2)
        self.sitesVisible=not self.sitesVisible
    def toggleTrash(self):
        if self.trashVisible:
            self.data_window.imageview.removeItem(self.s3)
        else:
            self.data_window.imageview.addItem(self.s3)
        self.trashVisible=not self.trashVisible
    
    def setCurrPuff(self,value,force=False):
        if force is False and self.currentPuff_spinbox.puffidx==value:
            return
        self.currentPuff_spinbox.puffidx=value
        self.puffs.setIndex(value)
        puff=self.puffs[value]
        self.threeD_plot.update_puff(puff)
        self.trace_plot.clear()
        puff.plot(self.trace_plot)
        self.trace_plot.plotItem.autoRange()
        
        if self.lastClicked is not None:
            self.lastClicked.resetPen()
        point=[s for s in self.s1.points() if s.data() is puff][0]
        point.setPen('r', width=2)
        # move the roi to that point
        pts=self.roi.getPoints()
        roi_pt=np.array([(pts[0][0]+pts[2][0])/2,(pts[0][1]+pts[2][1])/2])
        puff_pt=np.array([puff.kinetics['x'],puff.kinetics['y']])
        difference=puff_pt-roi_pt
        self.roi.beingDragged=True
        self.roi.translate(QPointF(*difference),QPointF(*roi_pt))
        self.roi.finish_translate()
        self.lastClicked = point
        if self.currentPuff_spinbox.value()!=self.puffs.index:
            self.currentPuff_spinbox.setValue(self.puffs.index)

    def clicked(self, plot, points):
        point=points[0]
        puff=point.data()
        self.setCurrPuff(self.puffs.puffs.index(puff))
    def clickedTrash(self,plot,points):
        point=points[0]
        puff=point.data()
        self.puffs.addPuffs([puff])
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        self.trash.removePuffs([puff])
        self.setCurrPuff(self.puffs.puffs.index(puff))
        
    def keyPressEvent(self,ev):
        nPuffs=len(self.puffs.puffs)
        old_idx=self.puffs.index
        if ev.key() == Qt.Key_Right:
            if old_idx==nPuffs-1:
                new_idx=0
            else:
                new_idx=old_idx+1
            self.setCurrPuff(new_idx)
        elif ev.key() == Qt.Key_Left:
            if old_idx==0:
                new_idx=nPuffs-1
            else:
                new_idx=old_idx-1
            self.setCurrPuff(new_idx)
        elif ev.key()==Qt.Key_Delete:
            if self.roi.mouseIsOver or self.roi.beingDragged:
                puffs=[pt.data() for pt in self.s1.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
                self.trash.addPuffs(puffs)
                self.sites.removePuffs(puffs)
                self.puffs.removePuffs(puffs)
                if len(self.puffs.puffs)==0:
                    self.close()
                    return 
                self.setCurrPuff(self.puffs.index,force=True)
                self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
                self.drawRedOverlay()                
            else: 
                #if np.any([self.roi.window.rois[i].mouseIsOver for i in np.arange(len(self.roi.window.rois))])==False: #if no ROIs are being hovered over with the mouse
                self.discard_currPuff()
        elif ev.key()==Qt.Key_G:
            #Group all events in roi
            sites=[pt.data() for pt in self.s2.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
            if len(sites)==0:
                return
            puffs=list(itertools.chain(*[site.puffs for site in sites]))
            for site in sites:
                self.sites.remove(site) 
            self.sites.append(Site(puffs))
            self.s2.clear()
            for site in self.sites:
                self.s2.addPoints(pos=[site.pos],data=site)
        elif ev.key()==Qt.Key_U:
            sites=[pt.data() for pt in self.s2.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
            if len(sites)==0:
                return
            puffs=list(itertools.chain(*[site.puffs for site in sites]))
            for site in sites:
                self.sites.remove(site) 
            self.sites.extend([Site([p]) for p in puffs])
            self.s2.clear()
            for site in self.sites:
                self.s2.addPoints(pos=[site.pos],data=site)
            
    def discard_currPuff(self):
        puff=self.puffs.getPuff()
        self.trash.addPuffs([puff])
        self.sites.removePuffs([puff])
        self.puffs.removePuffs([puff])
        self.setCurrPuff(self.puffs.index,force=True)
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        self.drawRedOverlay()

    def togglePuffs(self):
        if self.puffsVisible:
            self.data_window.imageview.removeItem(self.s1)
        else:
            self.data_window.imageview.addItem(self.s1)
        self.puffsVisible=not self.puffsVisible
        


        
    def drawRedOverlay(self):
        puffs=[pt.data() for pt in self.s1.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
        times=[[puff.kinetics['t_start'],puff.kinetics['t_end']] for puff in puffs]
        data=g.m.currentTrace.rois[g.m.currentTrace.get_roi_index(self.roi)]['p1trace'].getData()[1]
        
        x=np.array([np.arange(*times[i]) for i in np.arange(len(times))])
        traces=[data[time] for time in x]
        y=np.array(traces)
        for i in np.arange(len(self.redTraces)):
            g.m.currentTrace.p1.removeItem(self.redTraces[i][0])
        self.redTraces=[]
        for i in np.arange(len(x)):
            self.redTraces.append([g.m.currentTrace.p1.plot(x[i],y[i],pen=pg.mkPen('r')),puffs[i]])
        currentPuff=self.puffs.getPuff()
        if currentPuff in puffs:
            idx=puffs.index(currentPuff)
            self.redTraces[idx][0].setPen(pg.mkPen(color='r',width=3))


    def clickedTrace(self,ev):
        self.EEEE=ev
        pos=ev.pos()
        pos=g.m.currentTrace.vb.mapSceneToView(pos)
        t=pos.x()
        puffs=[pt.data() for pt in self.s1.points() if self.roi.contains(pt.pos().x(),pt.pos().y())]
        times=[[puff.kinetics['t_start'],puff.kinetics['t_end']] for puff in puffs]
        try:
            index=[(t<=time[1] and t>=time[0]) for time in times].index(True)
        except ValueError:
            return
        puff=puffs[index]
        if ev.button()==1:
            self.setCurrPuff(self.puffs.puffs.index(puff))
    def openFilterGUI(self):
        print('yay')
            
    def export_gui(self):
        filename=g.m.settings['filename']
        directory=os.path.dirname(filename)
        if filename is not None:
            filename= QFileDialog.getSaveFileName(g.m, 'Export Puff Info', directory, '*.xlsx')
        else:
            filename= QFileDialog.getSaveFileName(g.m, 'Export Puff Info', '*.xlsx')
        filename=str(filename)
        if filename=='':
            return False
        else:
            self.export(filename)
    def export(self,filename):
        ''' This function saves out all the info about the puffs        
        '''
        from openpyxl import Workbook
        from openpyxl.cell import get_column_letter
        g.m.statusBar().showMessage('Saving {}'.format(os.path.basename(filename)))
        
        workbook = Workbook() 
        sheet = workbook.create_sheet()
        sheet.title="Puff Data"
        header=['Group #','Group x','Group y','No. Events','Max Amp','x','y','t_peak','Amplitude','sigmax','sigmay','angle','r20','r50','r80','r100','f80','f50','f20','f0']
        for j in np.arange(len(header)):
            col = get_column_letter(j+1)
            sheet.cell("{}{}".format(col,1)).value=header[j]
        row=2
        groupN=1
        for site in self.sites:
            r=str(row)
            groupx,groupy=site.pos
            nEvents=len(site.puffs)
            maxAmp=np.max([puff.kinetics['amplitude'] for puff in site.puffs])
            sheet.cell('A'+r).value=groupN
            sheet.cell('B'+r).value=groupx
            sheet.cell('C'+r).value=groupy
            sheet.cell('D'+r).value=nEvents
            sheet.cell('E'+r).value=maxAmp
            for puff in site.puffs:
                r=str(row)
                k=puff.kinetics
                sheet.cell('F'+r).value=k['x']
                sheet.cell('G'+r).value=k['y']
                
                try: sheet.cell('H'+r).value=int(k['t_peak'])
                except ValueError: pass
                sheet.cell('I'+r).value=k['amplitude']
                if 'sigmax' in k.keys():
                    sheet.cell('J'+r).value=k['sigmax']
                    sheet.cell('K'+r).value=k['sigmay']
                    sheet.cell('L'+r).value=k['angle']
                else:
                    sheet.cell('J'+r).value=k['sigma']
                try: sheet.cell('M'+r).value=int(k['r20'])
                except ValueError: pass
                try: sheet.cell('N'+r).value=int(k['r50'])
                except ValueError: pass
                try: sheet.cell('O'+r).value=int(k['r80'])
                except ValueError: pass
                try: sheet.cell('P'+r).value=int(k['t_peak']-k['t_start'])
                except ValueError: pass
                try: sheet.cell('Q'+r).value=int(k['f80'])
                except ValueError: pass
                try: sheet.cell('R'+r).value=int(k['f50'])
                except ValueError: pass
                try: sheet.cell('S'+r).value=int(k['f20'])
                except ValueError: pass
                try: sheet.cell('T'+r).value=int(k['f0'])
                except ValueError: pass
                row+=1
            groupN+=1
            
        sheet = workbook.create_sheet()
        sheet.title="Site traces"
        groupN=1
        for site in self.sites:
            trace=self.data_window.image[:,site.pos[0],site.pos[1]]
            col=get_column_letter(groupN)
            sheet.cell(col+'1').value="Site #{}".format(groupN)
            for i in np.arange(len(trace)):
                sheet.cell(col+str(i+2)).value=trace[i]
            groupN+=1
        
        sheet = workbook.create_sheet()
        sheet.title="Peak aligned Puff Traces"
        siteN=1
        max_peak_idx=np.max([puff.kinetics['t_peak']-puff.kinetics['t_start'] for puff in self.puffs])
        for puff in self.puffs.puffs:
            col=get_column_letter(siteN)
            peak_idx=puff.kinetics['t_peak']-puff.kinetics['t_start']
            for i in np.arange(len(puff.trace)):
                offset=max_peak_idx-peak_idx
                sheet.cell(col+str(offset+i+1)).value=puff.trace[i]
            siteN+=1
        workbook.save(filename)
        g.m.statusBar().showMessage('Successfully saved {}'.format(os.path.basename(filename)))
            
        
    
def getPuffs(puffbool,data_im,weakfilt,strongfilt,paddingXY,paddingT_pre,paddingT_post,maxSigmaForGaussianFit,rotatedfit):
    s=ndimage.generate_binary_structure(3,2)
    lbl, num_features = ndimage.measurements.label(puffbool, structure=s)
    lbls = np.arange(1, num_features+1)
    
    def fn(val,pos):
        pos=[np.unravel_index(p,data_im.shape) for p in pos]
        pos=np.array(pos)
        p0=np.min(pos,0)
        p1=np.max(pos,0)
        return (p0,p1)
        
    bounds=ndimage.labeled_comprehension(data_im,lbl,lbls,fn,tuple,0,True)
    highpass_im=get_highpass_im(data_im)
    puffs=Puffs(bounds,data_im,highpass_im,weakfilt,strongfilt,paddingXY,paddingT_pre,paddingT_post,maxSigmaForGaussianFit,rotatedfit)
    return puffs
    
def get_highpass_im(data_image):
    b,a,padlen=butterworth_filter.makeButterFilter(1,.01,1)
    mx=data_image.shape[2]
    my=data_image.shape[1]
    newtif=np.zeros(data_image.shape)
    for i in np.arange(my):
        for j in np.arange(mx):
            newtif[:, i, j]=filtfilt(b,a, data_image[:, i, j], padlen=padlen)
    return newtif








class threeD_plot(gl.GLViewWidget):
    def __init__(self,puff,parent=None):
        super(threeD_plot,self).__init__(parent)
        self.setCameraPosition(distance=150,elevation=30,azimuth=90)
        image=np.copy(puff.gaussianFit)
        mmax=np.max(puff.mean_image)
        image=image/mmax
        self.p1 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        ##    red   = pow(z * colorMap[0] + colorMap[1], colorMap[2])
        ##    green = pow(z * colorMap[3] + colorMap[4], colorMap[5])
        ##    blue  = pow(z * colorMap[6] + colorMap[7], colorMap[8])
        self.p1.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p1.scale(1, 1, 15.0)
        self.p1.translate(-puff.paddingXY, -puff.paddingXY, 0)
        self.addItem(self.p1)
        
        image=np.copy(puff.mean_image)
        image=image/mmax
        self.p2 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        self.p2.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p2.scale(1, 1, 15.0)
        self.p2.translate(-puff.paddingXY, -puff.paddingXY, 0)
        self.addItem(self.p2)
        
        self.shiftx=int(np.ceil(image.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)
    def update_puff(self,puff):
        self.p1.translate(-self.shiftx,0,0)
        self.p2.translate(self.shiftx,0,0)
        image=np.copy(puff.gaussianFit)
        mmax=np.max(puff.mean_image)
        image=image/mmax
        self.p1.setData(z=image)
        image=np.copy(puff.mean_image)
        image=image/mmax
        self.p2.setData(z=image)
        self.shiftx=int(np.ceil(image.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)
        #self.setMinimumHeight(300)

class Sites(list):
    def __init__(self,puffAnalyzer):
        super(Sites,self).__init__()
        self.puffAnalyzer=puffAnalyzer
    def removePuffs(self,puffs):
        idxs=[]
        for puff in puffs:
            try:
                idx=[puff in self[i].puffs for i in np.arange(len(self))].index(True)
                idxs.append(idx)
            except ValueError:
                pass
        idxs=np.unique(idxs)
        idxs=list(idxs)
        idxs.sort(reverse=True)
        for i in idxs:
            site=self[i]
            scatterRemovePoints(self.puffAnalyzer.s2,[i])
            site.removePuffs(puffs)
            if len(site.puffs)==0:
                self.remove(site)
            else:
                scatterAddPoints(self.puffAnalyzer.s2,[site.pos],[site])
                #self.puffAnalyzer.s2.addPoints(pos=[site.pos],data=site)
            
class Site(object):
    def __init__(self,puffs):
        self.puffs=puffs
        self.pos=self.getPos()
    def getPos(self):
        x=np.mean(np.array([p.kinetics['x'] for p in self.puffs]))
        y=np.mean(np.array([p.kinetics['y'] for p in self.puffs]))
        return [x,y]
    def removePuffs(self,puffs):
        for puff in puffs:
            if puff in self.puffs:
                self.puffs.remove(puff)
        if len(self.puffs)==0:
            self.pos=[np.nan,np.nan]
        else:
            self.pos=self.getPos()
        
class Trash(list):
    def __init__(self,puffAnalyzer):
        super(Trash,self).__init__()
        self.puffAnalyzer=puffAnalyzer
    def addPuffs(self,puffs):
        self.extend(puffs)
        pos=[[puff.kinetics['x'],puff.kinetics['y']] for puff in puffs]
        #self.puffAnalyzer.s3.addPoints(pos=pos,data=puffs)
        scatterAddPoints(self.puffAnalyzer.s3,pos,puffs)
    def removePuffs(self,puffs):
        for puff in puffs:
            if puff in self:
                self.remove(puff)
        s=self.puffAnalyzer.s3
        idxs=[]
        for puff in puffs:
            idxs.append([point['data'] for point in s.data].index(puff))
        scatterRemovePoints(self.puffAnalyzer.s3,idxs)
    
class Puffs:
    def __init__(self,bounds,data_im,highpass_im,weakfilt,strongfilt,paddingXY,paddingT_pre,paddingT_post,maxSigmaForGaussianFit,rotatedfit):
        self.puffs=[]
        self.index=0
        self.data_im=data_im
        self.highpass_im=highpass_im
        self.weakfilt=weakfilt
        self.strongfilt=strongfilt
        self.paddingXY=paddingXY
        self.paddingT_pre=paddingT_pre
        self.paddingT_post=paddingT_post
        self.rotatedfit=rotatedfit
        self.maxSigmaForGaussianFit=maxSigmaForGaussianFit
        self.puffs=[Puff(bounds[i],self) for i in np.arange(len(bounds))]

    def __getitem__(self, item):
        return self.puffs[item]
    def removeCurrentPuff(self):
        del self.puffs[self.index]
        if self.index==0:
            return self.index
        else:
            self.index-=1
        return self.index
    def getPuff(self):
        return self.puffs[self.index]
    def increment(self):
        self.index+=1
        if len(self.puffs)<self.index+1:
            self.index=0
    def decrement(self):
        self.index-=1
        if self.index<0:
            self.index=len(self.puffs)-1
    def setIndex(self,index):
        self.index=index
        if len(self.puffs)<self.index+1:
            self.index=0
        elif self.index<0:
            self.index=len(self.puffs)-1
    def removePuffs(self,puffs):
        idxs=[]
        for puff in puffs:
            idxs.append([point['data'] for point in self.puffAnalyzer.s1.data].index(puff))
            self.puffs.remove(puff)
        scatterRemovePoints(self.puffAnalyzer.s1,idxs)
        if self.index>=len(self.puffs):
            self.index=len(self.puffs)-1
    def addPuffs(self,puffs):
        s=self.puffAnalyzer.s1
        self.puffs.extend(puffs)
        pos=[[puff.kinetics['x'],puff.kinetics['y']] for puff in puffs]
        scatterAddPoints(s,pos,puffs)
        #s.addPoints(pos=pos,data=puffs)

        
class Puff:
    def __init__(self,originalbounds,puffs):
        originalbounds=np.array(originalbounds) # 2x3 array: [[t_min,x_min,y_min],[t_max,x_max,y_max]]
        self.puffs=puffs
        self.paddingXY=self.puffs.paddingXY
        self.originalbounds=originalbounds
        self.trace=None
        self.kinetics=dict()
        
        #######################################################################
        #############          FIND (x,y) ORIGIN       ########################
        #######################################################################
        '''
        For debugging, use the following code:
        self=g.m.puffAnalyzer.puffs.getPuff()
        from analyze.puffs.average_origin import *
        '''
        
        t0=self.originalbounds[0][0]
        t1=self.originalbounds[1][0]
        x0=self.originalbounds[0][1]-self.puffs.paddingXY
        x1=self.originalbounds[1][1]+self.puffs.paddingXY
        y0=self.originalbounds[0][2]-self.puffs.paddingXY
        y1=self.originalbounds[1][2]+self.puffs.paddingXY
        mt,mx,my=self.puffs.highpass_im.shape
        if t0<0: t0=0
        if y0<0: y0=0
        if x0<0: x0=0
        if t1>=mt: t1=mt-1
        if y1>=my: y1=my-1
        if x1>=mx: x1=mx-1
        self.bounds=[(t0,t1),(x0,x1),(y0,y1)]
        bb=self.bounds
        I=self.puffs.highpass_im[bb[0][0]:bb[0][1]+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        I=np.mean(I,0)
        xorigin,yorigin=(np.array(I.shape)/2).astype(np.int)
        sigma=3
        amplitude=np.max(I)/2
        if self.puffs.rotatedfit:
            sigmax=sigma
            sigmay=sigma
            angle=45
            p0=(xorigin,yorigin,sigmax,sigmay,angle,amplitude)
            #                 xorigin                   yorigin             sigmax, sigmay, angle,    amplitude
            fit_bounds = [(0.0, float(I.shape[0])), (0, float(I.shape[1])),  (2,self.puffs.maxSigmaForGaussianFit), (2,self.puffs.maxSigmaForGaussianFit), (0,90),   (0,np.max(I))]
            p, I_fit, I2= fitRotGaussian(I,p0,fit_bounds)
            self.mean_image=I
            self.gaussianFit=I_fit
            p[0]=p[0]+self.bounds[1][0] #Put back in regular coordinate system.  Add back x
            p[1]=p[1]+self.bounds[2][0] #add back y 
            self.gaussianParams=p
            xorigin,yorigin,sigmax,sigmay,angle,amplitude=self.gaussianParams
        else:
            p0=(xorigin,yorigin,sigma,amplitude) 
            #                 xorigin                   yorigin            sigma    amplitude
            fit_bounds = [(0.0, float(I.shape[0])), (0, float(I.shape[1])),  (2,self.puffs.maxSigmaForGaussianFit),    (0,np.max(I))] #[(0.0, 2*self.paddingXY), (0, 2*self.paddingXY),(0,10),(0,10),(0,90),(0,5)]
            p, I_fit, I2= fitGaussian(I,p0,fit_bounds)
            self.mean_image=I
            self.gaussianFit=I_fit
            p[0]=p[0]+self.bounds[1][0] #Put back in regular coordinate system.  Add back x
            p[1]=p[1]+self.bounds[2][0] #add back y 
            self.gaussianParams=p
            xorigin,yorigin,sigma,amplitude=self.gaussianParams

        
        if self.puffs.rotatedfit:
            self.kinetics['sigmax']=sigmax; self.kinetics['sigmay']=sigmay; self.kinetics['angle']=angle
        else:
            self.kinetics['sigma']=sigma;
        self.kinetics['x']=xorigin; self.kinetics['y']=yorigin;
        #######################################################################
        #############          FIND PEAK       ########################
        #######################################################################
        if amplitude==0:
            I_norm=np.zeros(self.gaussianFit.shape)
        else:
            I_norm=I_fit/np.sum(I_fit)
        before=bb[0][0]-self.puffs.paddingT_pre
        after=bb[0][1]+self.puffs.paddingT_post
        if before<0: before=0
        if after>=mt: after=mt-1;
        self.kinetics['before']=before
        self.kinetics['after']=after
        I=self.puffs.highpass_im[before:after+1,bb[1][0]:bb[1][1]+1,bb[2][0]:bb[2][1]+1]
        trace=np.zeros((len(I)))
        for i in np.arange(len(trace)):
            trace[i]=2*np.sum(I[i]*I_norm)+1
            #I'm not sure why the 2 is needed, but this seems to always work out to 1:
            #            from analyze.puffs.gaussianFitting import gaussian
            #            x = np.arange(1000,dtype=float)
            #            y = np.arange(1000,dtype=float)
            #            amplitude=2.8
            #            sigma=7
            #            I=gaussian(x[:,None], y[None,:],100,100,sigma,amplitude)
            #            I_fit=gaussian(x[:,None], y[None,:],100,100,sigma,1)
            #            I_norm=I_fit/np.sum(I_fit)
            #            calc_amp=2*np.sum(I*I_norm)
            #            print('Original amp={}  Calculated amp={}'.format(amplitude,calc_amp))
        self.trace=trace
        try:
            b1,a1,padlen1=butterworth_filter.makeButterFilter(1,0,self.puffs.strongfilt) #strong lowpass filter, plotted in green
            b2,a2,padlen2=butterworth_filter.makeButterFilter(1,0,self.puffs.weakfilt)  #weak  lowpass filter, plotted in red
            fStrong=filtfilt(b1,a1,trace,padlen=padlen1)
            fWeak=filtfilt(b2,a2,trace,padlen=padlen2)
        except ValueError: #this occurs when the puff is closer to the beginning or end than the padlen1 or padlen2
            fStrong=trace
            fWeak=trace
        

        self.fStrong=fStrong
        self.fWeak=fWeak

        
        # use maximum slope to find the rising phase of the puff.
        # It must also be during when an event was detected
        f_prime=np.insert(np.diff(fStrong),0,0)
        detected_times=[bb[0][0]-before, bb[0][1]-before]
        detected_times=np.arange(detected_times[0],detected_times[1]+1)
        tmp=np.zeros(len(f_prime))
        tmp[detected_times]=1
        tmp=f_prime*tmp
        if np.sum(tmp>0)==0: #if all slopes are less than zero
            tmp2=np.argwhere(f_prime>0)
            max_slope_idx=tmp2[np.abs(tmp2-np.mean(detected_times)).argmin()][0] #use the nearest positive slope
        else:
            max_slope_idx=np.argmax(tmp)
        max_slope=f_prime[max_slope_idx]
        maxima,minima=findextrema(fStrong)
        rising_times=np.where(f_prime>.1*max_slope)[0]
        rising_times=rising_times[rising_times<max_slope_idx]
        
        falling_times=np.where(f_prime<0)[0]
        falling_times=falling_times[falling_times<max_slope_idx]
        if len(falling_times)>0:
            rising_times=rising_times[rising_times>np.max(falling_times)]
            
        if len(rising_times)==0:
            t_start=0
        else:
            t_start=np.min(rising_times)-1
        
        baseline=fStrong[t_start]
        tmp=np.where(fWeak>baseline)[0]; tmp=tmp[np.logical_and(tmp>=t_start,tmp<=max_slope_idx)] # this finds indecies where fWeak is above baseline before the max slope and after 't_start'
        if len(tmp)==0:
            if max_slope_idx==0:
                t_start=0
            else:
                t_start=max_slope_idx-1
        else:
            if tmp[0]==0:
                t_start=0
            else:
                t_start=tmp[0]-1
        
        # use the next local maxima of fStrong as the time of the peak
        maximum=maxima[maxima>=max_slope_idx]
        if len(maximum)>0:
            t_peak=np.min(maximum)
        else:
            t_peak=max_slope_idx
        # stop looking for a decrease in puff the next time the slope is 10% of the value of the maximum slope
        falling_times=np.where(f_prime<-.1*max_slope)[0]
        falling_times=falling_times[falling_times>t_peak]
        rising_times=np.where(f_prime>.1*max_slope)[0]
        rising_times=rising_times[rising_times>t_peak]
        if len(rising_times)>0:
            falling_times=falling_times[falling_times<np.min(rising_times)]
        self.falling_times=falling_times
        self.kinetics['t_start']=t_start+before
        self.calcRiseFallTimes()
        
    def calcRiseFallTimes(self):
        before=self.kinetics['before']
        t_start=self.kinetics['t_start']-before
        if len(self.falling_times)>0:
            t_end=np.max(self.falling_times)
        else:
            t_end=len(self.fStrong)
        baseline=self.fStrong[t_start]
        t_peak=np.argmax(self.fWeak[t_start:t_end])+t_start  # find a more accurate t_peak
        f_peak=self.fWeak[t_peak]
        if baseline>f_peak:
            baseline=self.fWeak[t_start]
        amplitude=f_peak-baseline
        thresh20=baseline+amplitude*.2
        thresh50=baseline+amplitude*.5
        thresh80=baseline+amplitude*.8
        tmp=np.argwhere(self.fWeak>thresh20); tmp=tmp[np.logical_and(tmp>=t_start,tmp<=t_peak)]; 
        if len(tmp)==0: r20=np.nan
        else:  r20=tmp[0]-t_start
        tmp=np.argwhere(self.fWeak>thresh50); tmp=tmp[np.logical_and(tmp>=t_start,tmp<=t_peak)];
        if len(tmp)==0: r50=np.nan
        else:  r50=tmp[0]-t_start
        tmp=np.argwhere(self.fWeak>thresh80); tmp=tmp[np.logical_and(tmp>=t_start,tmp<=t_peak)]; 
        if len(tmp)==0: r80=np.nan
        else:  r80=tmp[0]-t_start
        tmp=np.argwhere(self.fWeak<thresh80); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f80=np.nan
        else: f80=tmp[0]-t_peak
        
        tmp=np.argwhere(self.fWeak<thresh50); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f50=np.nan
        else: f50=tmp[0]-t_peak
        
        tmp=np.argwhere(self.fWeak<thresh20); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f20=np.nan
        else: f20=tmp[0]-t_peak
        
        tmp=np.argwhere(self.fWeak<baseline); tmp=tmp[tmp>=t_peak]; 
        if len(tmp)==0: f0=np.nan
        else: 
            f0=tmp[0]
            if f0<t_end:
                t_end=f0
            f0=f0-t_peak
            
        self.kinetics['amplitude']=amplitude
        self.kinetics['baseline']=baseline
        self.kinetics['t_end']=t_end+before
        self.kinetics['r20']=r20
        self.kinetics['r50']=r50
        self.kinetics['r80']=r80
        self.kinetics['f20']=f20
        self.kinetics['f50']=f50
        self.kinetics['f80']=f80
        self.kinetics['f0']=f0
        self.kinetics['t_peak']=t_peak+before

            
    def plot(self,figure=None):
        if figure is None:
            figure=pg.plot()
        k=self.kinetics
        baseline=k['baseline']; amplitude=k['amplitude']
        #thresh20=baseline+amplitude*.2
        #thresh50=baseline+amplitude*.5
        #thresh80=baseline+amplitude*.8
        x=np.arange(len(self.trace))+k['before']
        figure.plot(x,self.trace,pen=pg.mkPen(width=2))
        figure.plot(x,self.fStrong,pen=pg.mkPen('g'))
        figure.plot(x,self.fWeak,pen=pg.mkPen('r'))
        self.peakLine=figure.addLine(y=baseline,pen=pg.mkPen('y',style=Qt.DashLine))
        self.baselineLine=figure.addLine(y=baseline+amplitude,pen=pg.mkPen('y',style=Qt.DashLine))
        self.startLine=figure.addLine(x=k['t_start'],pen=pg.mkPen('y',style=Qt.DashLine),movable=True,bounds=(self.kinetics['before'],self.kinetics['t_peak']))
        self.endLine=figure.addLine(x=k['t_end'],pen=pg.mkPen('y',style=Qt.DashLine),movable=True, bounds=(self.kinetics['t_peak'],self.kinetics['after']))
        self.startLine.sigDragged.connect(self.changeStartTime)
        self.endLine.sigDragged.connect(self.changeEndTime)
    def changeStartTime(self,line):
        time=line.value()
        time=int(np.round(time))
        if time!=line.value():
            self.startLine.setValue(time)
        oldstart=self.kinetics['t_start']
        self.kinetics['t_start']=time
        if oldstart!=time:
            self.calcRiseFallTimes()
            self.baselineLine.setValue(self.kinetics['baseline'])
            self.peakLine.setValue(self.kinetics['baseline']+self.kinetics['amplitude'])
            self.endLine.setValue(self.kinetics['t_end'])
    def changeEndTime(self,line):
        time=line.value()
        time=int(np.round(time))
        if time!=line.value():
            self.endLine.setValue(time)
        oldend=self.kinetics['t_end']
        if oldend!=time:   
            self.puffs.puffAnalyzer.drawRedOverlay()
            self.kinetics['t_end']=time
        
        



def findextrema(v):
    v_prime=np.concatenate(([0], np.diff(v)))
    zero_crossing=np.concatenate((np.diff(np.sign(v_prime)),[0]))
    maxima=[1 if v < -1 else 0 for v in zero_crossing]
    minima=[1 if v > 1 else 0 for v in zero_crossing]
    if np.sign(v_prime[-1])>0:
        maxima[-1]=1
    else:
        minima[-1]=1;
    if np.sign(v_prime[0])>0:
        maxima[0]=1
    else:
        minima[0]=1
    maxima=np.nonzero(maxima)[0]
    minima=np.nonzero(minima)[0]
    return maxima, minima



def scatterRemovePoints(scatterplot,idxs):
    i2=[i for i in np.arange(len(scatterplot.data)) if i not in idxs]
    points=scatterplot.points()
    points=points[i2]
    spots=[{'pos':points[i].pos(),'data':points[i].data(),'brush':points[i].brush()} for i in np.arange(len(points))]
    #spots=[{'pos':points[i].pos(),'data':points[i].data(),'brush':points[i].brush(),'pen':points[i].pen()} for i in np.arange(len(points))]
    scatterplot.clear()
    scatterplot.addPoints(spots)
def scatterAddPoints(scatterplot,pos,data):
    points=scatterplot.points()
    spots=[{'pos':points[i].pos(),'data':points[i].data()} for i in np.arange(len(points))]
    spots.extend([{'pos':pos[i],'data':data[i]} for i in np.arange(len(pos))])
    scatterplot.clear()
    scatterplot.addPoints(spots)
#    scatterplot.data=np.empty(len(oldData)-numPts,dtype=scatterplot.data.dtype)
#    i2=[i for i in np.arange(len(oldData)) if i not in idxs]
#    scatterplot.data=oldData[i2]
#    scatterplot.prepareGeometryChange()
#    scatterplot.bounds = [None, None]
#    scatterplot.invalidate()
#    scatterplot.updateSpots(oldData)
#    scatterplot.sigPlotChanged.emit(scatterplot)
#    for pt in scatterplot.data:
#        if pt['item'] is not None:
#            pt['item']._data=pt





